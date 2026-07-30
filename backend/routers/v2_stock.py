from fastapi import APIRouter, Depends, HTTPException, UploadFile, File, Form, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session, joinedload
from typing import List, Optional
from ..database import get_db
from .. import models, schemas
from ..core.security import (
    assert_permission,
    get_current_user,
    roles_have_permission,
)
import time
import io
import tempfile
from pathlib import Path
from datetime import datetime, timedelta, timezone
from uuid import uuid4
import openpyxl
from openpyxl.styles import Font, PatternFill
from sqlalchemy import func, or_, update
from sqlalchemy.exc import IntegrityError
from ..services.bom_parser import parse_bom_file
from ..services.stock_reservations import (
    InsufficientStockAtConsumptionError,
    annotate_variant_availability,
    available_quantity_at_location,
    build_preview_payload,
    cancel_reservation,
    consume_reservation,
    create_reservation,
    technical_dossier_for_sale,
)
from ..services.stock_service import InventoryService
from ..services.workshop_preparations import (
    cancel_preparation,
    create_preparation,
    hand_over_preparation,
    load_preparation,
    return_preparation,
    update_prepared_quantity,
)
from scripts.import_workshop_debits import parse_file
from ..core.time import utcnow

router = APIRouter(
    prefix="/v2/stock",
    tags=["stock"],
    dependencies=[Depends(get_current_user)],
)

CATALOG_STATUSES = {"DRAFT", "TO_QUALIFY", "ACTIVE", "BLOCKED", "ARCHIVED"}
CATALOG_TRANSITIONS = {
    "DRAFT": {"TO_QUALIFY", "ARCHIVED"},
    "TO_QUALIFY": {"DRAFT", "ACTIVE", "ARCHIVED"},
    "ACTIVE": {"BLOCKED", "ARCHIVED"},
    "BLOCKED": {"ACTIVE", "ARCHIVED"},
    "ARCHIVED": {"DRAFT"},
}


def _actor(user: dict) -> str:
    return str(user.get("sub") or user.get("username") or "system")


def _clean(value: Optional[str]) -> Optional[str]:
    cleaned = str(value or "").strip()
    return cleaned or None


def _record_product_audit(
    db: Session,
    *,
    product_id: int,
    user: dict,
    action: str,
    changes: Optional[dict] = None,
    variant_id: Optional[int] = None,
    reason: Optional[str] = None,
) -> None:
    db.add(
        models.ProductAuditLog(
            product_id=product_id,
            variant_id=variant_id,
            action=action,
            changes=changes or None,
            reason=_clean(reason),
            author=_actor(user),
            created_at=utcnow(),
        )
    )


def _json_value(value):
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value)


def _changes(instance, values: dict) -> dict:
    result = {}
    for key, new_value in values.items():
        old_value = getattr(instance, key, None)
        if old_value != new_value:
            result[key] = {"before": _json_value(old_value), "after": _json_value(new_value)}
    return result


def _ensure_variant_uniqueness(
    db: Session,
    *,
    reference: str,
    barcode: Optional[str],
    supplier_reference: Optional[str],
    supplier: Optional[str],
    exclude_variant_id: Optional[int] = None,
) -> None:
    reference_query = db.query(models.ProductVariant).filter(
        func.upper(models.ProductVariant.reference) == reference.upper()
    )
    if exclude_variant_id:
        reference_query = reference_query.filter(models.ProductVariant.id != exclude_variant_id)
    if reference_query.first():
        raise HTTPException(409, f"La référence interne {reference} existe déjà.")

    if barcode:
        barcode_query = db.query(models.ProductVariant).filter(models.ProductVariant.barcode == barcode)
        if exclude_variant_id:
            barcode_query = barcode_query.filter(models.ProductVariant.id != exclude_variant_id)
        if barcode_query.first():
            raise HTTPException(409, f"Le code-barres {barcode} est déjà utilisé.")

    if supplier and supplier_reference:
        supplier_query = (
            db.query(models.ProductVariant)
            .join(models.Product, models.ProductVariant.product_id == models.Product.id)
            .filter(
                func.upper(models.Product.supplier) == supplier.upper(),
                func.upper(models.ProductVariant.supplier_reference) == supplier_reference.upper(),
            )
        )
        if exclude_variant_id:
            supplier_query = supplier_query.filter(models.ProductVariant.id != exclude_variant_id)
        if supplier_query.first():
            raise HTTPException(
                409,
                f"La référence fournisseur {supplier_reference} existe déjà pour {supplier}.",
            )


def _activation_issues(product: models.Product) -> list[str]:
    issues = []
    for label, value in (
        ("désignation", product.name),
        ("référence famille", product.reference_base),
        ("catégorie", product.category),
        ("unité de gestion", product.unit),
    ):
        if not _clean(value):
            issues.append(label)

    if product.product_type != "service":
        if not _clean(product.material_type):
            issues.append("matière")
        if not _clean(product.supplier):
            issues.append("fournisseur principal")
        if not product.variants:
            issues.append("au moins une variante")
        for variant in product.variants:
            if not _clean(variant.reference):
                issues.append("référence interne de variante")
            if not _clean(variant.supplier_reference):
                issues.append(f"référence fournisseur de {variant.reference or 'la variante'}")
            if product.unit == "barre" and not (variant.length_per_unit and variant.length_per_unit > 0):
                issues.append(f"longueur de {variant.reference or 'la variante'}")
    elif not product.variants:
        issues.append("au moins une tarification de prestation")
    return list(dict.fromkeys(issues))


async def _parse_workshop_uploads(files: List[UploadFile]):
    records = []
    issues = []
    source_names = []
    for file in files:
        suffix = Path(file.filename or "").suffix
        with tempfile.NamedTemporaryFile(delete=False, suffix=suffix) as tmp:
            tmp.write(await file.read())
            tmp_path = Path(tmp.name)
        try:
            parsed_records, parsed_issues = parse_file(tmp_path)
            source_name = file.filename or tmp_path.name
            for record in parsed_records:
                object.__setattr__(record, "source", source_name)
            for issue in parsed_issues:
                object.__setattr__(issue, "source", source_name)
            records.extend(parsed_records)
            issues.extend(parsed_issues)
            source_names.append(source_name)
        finally:
            tmp_path.unlink(missing_ok=True)
    return records, issues, source_names


def _require_permission(db: Session, user: dict, permission_code: str) -> None:
    assert_permission(db, user, permission_code)


def _has_permission(db: Session, user: dict, permission_code: str) -> bool:
    try:
        assert_permission(db, user, permission_code)
        return True
    except HTTPException as exc:
        if exc.status_code in {401, 403}:
            return False
        raise


def _user_roles(user: dict) -> set[str]:
    roles = set(user.get("roles") or [])
    if user.get("role"):
        roles.add(user["role"])
    return roles


def _normalize_usernames(values: Optional[List[str]]) -> List[str]:
    return list(dict.fromkeys(
        str(value or "").strip()
        for value in (values or [])
        if str(value or "").strip()
    ))


def _assert_known_inventory_counters(db: Session, usernames: List[str]) -> None:
    if not usernames:
        return
    rows = (
        db.query(models.User)
        .filter(models.User.username.in_(usernames), models.User.is_active == True)
        .all()
    )
    found = {row.username for row in rows}
    missing = [username for username in usernames if username not in found]
    if missing:
        raise HTTPException(
            status_code=400,
            detail=f"Compteur(s) actif(s) introuvable(s): {', '.join(missing)}",
        )
    unauthorized = [
        row.username
        for row in rows
        if not roles_have_permission(db, row.role_names, "inventory.count")
    ]
    if unauthorized:
        raise HTTPException(
            status_code=400,
            detail=(
                "Permission inventory.count absente pour: "
                f"{', '.join(unauthorized)}"
            ),
        )


def _assert_counter_assignment(session: models.InventorySession, user: dict) -> None:
    assigned = set(session.assigned_usernames or [])
    if not assigned:
        return
    username = _actor(user)
    if username in assigned:
        return
    if _user_roles(user) & {"ADMIN", "SUPER_ADMIN", "MANAGER"}:
        return
    raise HTTPException(
        status_code=403,
        detail=f"Vous n'êtes pas affecté à la campagne {session.reference}.",
    )


def _active_blind_location_ids(db: Session, user: dict) -> set[int]:
    """Emplacements dont l'espéré doit être masqué pour le compteur courant.

    Les valideurs conservent leur vue de contrôle. Un compteur affecté (ou tous
    les compteurs lorsqu'aucune affectation n'est définie) ne reçoit ni quants
    couverts, ni agrégats variante tant que la campagne aveugle est ouverte.
    """
    if not _has_permission(db, user, "inventory.count"):
        return set()
    if _has_permission(db, user, "inventory.validate"):
        return set()
    username = _actor(user)
    sessions = (
        db.query(models.InventorySession)
        .filter(
            models.InventorySession.blind_counting == True,
            models.InventorySession.archived_at == None,
            models.InventorySession.status.in_(["draft", "counting", "pending_approval"]),
        )
        .all()
    )
    covered: set[int] = set()
    for session in sessions:
        assigned = set(session.assigned_usernames or [])
        if assigned and username not in assigned:
            continue
        covered.update(_zone_location_ids(db, session.location_id))
    return covered


def _can_reveal_blind_inventory(db: Session, user: dict) -> bool:
    return _has_permission(db, user, "inventory.approve_value")


def _get_quant_quantity(db: Session, variant_id: int, location_id: int) -> float:
    quant = db.query(models.StockQuant).filter_by(variant_id=variant_id, location_id=location_id).first()
    return float(quant.quantity if quant else 0)


def _get_or_create_quant(db: Session, variant_id: int, location_id: int) -> models.StockQuant:
    quant = db.query(models.StockQuant).filter_by(variant_id=variant_id, location_id=location_id).first()
    if not quant:
        quant = models.StockQuant(variant_id=variant_id, location_id=location_id, quantity=0)
        db.add(quant)
        db.flush()
    return quant


def _get_or_create_inventory_location(db: Session) -> models.StockLocation:
    location = db.query(models.StockLocation).filter_by(name="Virtual/Inventory", usage="inventory").first()
    if not location:
        location = models.StockLocation(name="Virtual/Inventory", usage="inventory", is_active=True)
        db.add(location)
        db.flush()
    return location


def _line_status_from_variance(variance: float) -> str:
    return "ok" if abs(float(variance or 0)) <= 0.000001 else "variance"


def _line_variance_value(line: models.InventoryCountLine) -> float:
    if line.variance_value is not None:
        return float(line.variance_value)
    return float(line.variance_quantity or 0) * float(line.unit_cost_snapshot or 0)


def _inventory_value_summary(session: models.InventorySession) -> tuple[float, float]:
    values = [_line_variance_value(line) for line in (session.lines or [])]
    return sum(values), sum(abs(value) for value in values)


def _naive_utc(value: Optional[datetime]) -> Optional[datetime]:
    if value is None:
        return None
    if value.tzinfo is None:
        return value
    return value.astimezone(timezone.utc).replace(tzinfo=None)


def _zone_location_ids(db: Session, location_id: Optional[int]) -> List[int]:
    """Emplacements de la zone d'inventaire : cible + descendants, ou tous les
    emplacements internes actifs pour une campagne globale."""
    if location_id is None:
        rows = (
            db.query(models.StockLocation.id)
            .filter(models.StockLocation.usage == "internal", models.StockLocation.is_active == True)
            .all()
        )
        return [row.id for row in rows]
    ids = [location_id]
    frontier = [location_id]
    while frontier:
        children = (
            db.query(models.StockLocation.id)
            .filter(models.StockLocation.parent_id.in_(frontier))
            .all()
        )
        new_ids = [row.id for row in children if row.id not in ids]
        ids.extend(new_ids)
        frontier = new_ids
    return ids


def _prefill_inventory_lines(db: Session, session: models.InventorySession, include_all_variants: bool) -> None:
    """Pré-remplit les lignes de comptage (statut ``pending``) depuis les quants
    de la zone, expected figé à la création. ``include_all_variants`` ajoute les
    variantes actives sans stock dans la zone (espéré 0) pour détecter les oublis."""
    zone_ids = _zone_location_ids(db, session.location_id)
    if not zone_ids:
        return
    quants = (
        db.query(models.StockQuant)
        .filter(models.StockQuant.location_id.in_(zone_ids))
        .order_by(models.StockQuant.variant_id, models.StockQuant.location_id)
        .all()
    )
    covered_variant_ids = set()
    for quant in quants:
        covered_variant_ids.add(quant.variant_id)
        db.add(models.InventoryCountLine(
            session_id=session.id,
            variant_id=quant.variant_id,
            location_id=quant.location_id,
            expected_quantity=float(quant.quantity or 0),
            status="pending",
        ))
    if include_all_variants:
        # Les variantes sans stock dans la zone sont ancrées sur l'emplacement
        # cible (ou le premier emplacement interne pour une campagne globale) :
        # l'opérateur constate explicitement le 0 ou corrige l'emplacement au comptage.
        anchor_location_id = session.location_id or min(zone_ids)
        active_variants = (
            db.query(models.ProductVariant)
            .join(models.Product, models.ProductVariant.product_id == models.Product.id)
            .filter(models.Product.catalog_status == "ACTIVE")
            .order_by(models.ProductVariant.id)
            .all()
        )
        for variant in active_variants:
            if variant.id in covered_variant_ids:
                continue
            db.add(models.InventoryCountLine(
                session_id=session.id,
                variant_id=variant.id,
                location_id=anchor_location_id,
                expected_quantity=0.0,
                status="pending",
            ))


def _mask_pending_line(line: schemas.InventoryCountLineResponse) -> None:
    """Une ligne ``pending`` n'a pas encore été comptée : le compté/opérateur
    ne sont pas exposés (la colonne garde 0 par défaut SQLAlchemy)."""
    if line.status == "pending":
        line.counted_quantity = None
        line.variance_quantity = None
        line.counted_by = None
        line.counted_at = None


def _serialize_inventory_session(
    session: models.InventorySession,
    *,
    reveal_blind: bool = False,
) -> schemas.InventorySessionResponse:
    """Sérialise une campagne ; en comptage aveugle, l'espéré et l'écart des
    lignes sont masqués tant que la campagne n'est pas validée/annulée."""
    response = schemas.InventorySessionResponse.model_validate(session)
    blind = (
        session.blind_counting
        and session.status in ["draft", "counting", "pending_approval"]
        and not (reveal_blind and session.status == "pending_approval")
    )
    total_value, absolute_value = _inventory_value_summary(session)
    threshold = float(session.approval_threshold_value or 0)
    response.total_variance_value = None if blind else total_value
    response.absolute_variance_value = None if blind else absolute_value
    response.requires_finance_approval = bool(
        threshold > 0
        and absolute_value > threshold + 0.005
        and not session.finance_approved_at
    )
    response.can_view_expected = not blind
    attachment_paths = {
        attachment.id: (
            f"/uploads/inventory/{session.id}/{model_line.id}/{attachment.stored_filename}"
        )
        for model_line in (session.lines or [])
        for attachment in (model_line.attachments or [])
    }
    for line in response.lines:
        _mask_pending_line(line)
        for attachment in line.attachments:
            attachment.url = attachment_paths.get(attachment.id)
        if blind:
            line.expected_quantity = None
            line.variance_quantity = None
            line.unit_cost_snapshot = None
            line.variance_value = None
            if line.status in {"ok", "variance"}:
                line.status = "counted"
    return response


def _serialize_count_line(
    session: models.InventorySession,
    line: models.InventoryCountLine,
    *,
    reveal_blind: bool = False,
) -> schemas.InventoryCountLineResponse:
    response = schemas.InventoryCountLineResponse.model_validate(line)
    _mask_pending_line(response)
    for attachment, row in zip(response.attachments, line.attachments):
        attachment.url = f"/uploads/inventory/{session.id}/{line.id}/{row.stored_filename}"
    if (
        session.blind_counting
        and session.status in ["draft", "counting", "pending_approval"]
        and not (reveal_blind and session.status == "pending_approval")
    ):
        response.expected_quantity = None
        response.variance_quantity = None
        response.unit_cost_snapshot = None
        response.variance_value = None
        if response.status in {"ok", "variance"}:
            response.status = "counted"
    return response


# Le gel de zone des campagnes d'inventaire est appliqué au niveau du moteur
# de mouvements : ``InventoryService.assert_location_not_locked`` (appelé par
# ``InventoryService.move_stock``) remonte la chaîne ``parent_id`` et bloque
# tout mouvement sur un emplacement couvert par une campagne gelée ouverte,
# y compris via un ancêtre. Ne pas réintroduire de duplicata local ici.


def _sync_variant_internal_stock(db: Session, variant_id: int) -> None:
    variant = db.query(models.ProductVariant).filter_by(id=variant_id).first()
    if not variant:
        return
    internal_quants = (
        db.query(models.StockQuant)
        .join(models.StockLocation, models.StockQuant.location_id == models.StockLocation.id)
        .filter(
            models.StockQuant.variant_id == variant_id,
            models.StockLocation.usage == "internal",
            models.StockLocation.is_active == True,
        )
        .all()
    )
    variant.quantity_in_stock = sum(float(quant.quantity or 0) for quant in internal_quants)


def _inventory_session_query(db: Session):
    return db.query(models.InventorySession).options(
        joinedload(models.InventorySession.location),
        joinedload(models.InventorySession.lines)
        .joinedload(models.InventoryCountLine.variant)
        .joinedload(models.ProductVariant.product),
        joinedload(models.InventorySession.lines)
        .joinedload(models.InventoryCountLine.location),
        joinedload(models.InventorySession.lines)
        .joinedload(models.InventoryCountLine.attachments),
    )


def _filtered_inventory_sessions_query(
    db: Session,
    *,
    status: Optional[str],
    search: Optional[str],
    include_archived: bool,
):
    query = _inventory_session_query(db)
    if status:
        query = query.filter(models.InventorySession.status == status)
    if not include_archived:
        query = query.filter(models.InventorySession.archived_at == None)
    cleaned_search = _clean(search)
    if cleaned_search:
        pattern = f"%{cleaned_search}%"
        query = query.filter(or_(
            models.InventorySession.reference.ilike(pattern),
            models.InventorySession.name.ilike(pattern),
            models.InventorySession.created_by.ilike(pattern),
        ))
    return query


@router.get("/inventory-sessions", response_model=List[schemas.InventorySessionResponse])
def list_inventory_sessions(
    status: Optional[str] = None,
    search: Optional[str] = None,
    include_archived: bool = False,
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    query = _filtered_inventory_sessions_query(
        db,
        status=status,
        search=search,
        include_archived=include_archived,
    )
    sessions = (
        query.order_by(models.InventorySession.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    reveal_blind = _can_reveal_blind_inventory(db, user)
    return [
        _serialize_inventory_session(session, reveal_blind=reveal_blind)
        for session in sessions
    ]


@router.get("/inventory-counters", response_model=List[schemas.User])
def list_inventory_counters(
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    users = (
        db.query(models.User)
        .filter(models.User.is_active == True)
        .order_by(
            models.User.first_name,
            models.User.last_name,
            models.User.username,
        )
        .all()
    )
    return [
        candidate
        for candidate in users
        if roles_have_permission(
            db,
            candidate.role_names,
            "inventory.count",
        )
    ]


@router.get("/inventory-sessions-page", response_model=schemas.InventorySessionPage)
def page_inventory_sessions(
    status: Optional[str] = None,
    search: Optional[str] = None,
    include_archived: bool = False,
    limit: int = Query(default=25, ge=1, le=100),
    offset: int = Query(default=0, ge=0),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    query = _filtered_inventory_sessions_query(
        db,
        status=status,
        search=search,
        include_archived=include_archived,
    )
    total = query.order_by(None).count()
    sessions = (
        query.order_by(models.InventorySession.created_at.desc())
        .offset(offset)
        .limit(limit)
        .all()
    )
    reveal_blind = _can_reveal_blind_inventory(db, user)
    return {
        "items": [
            _serialize_inventory_session(
                session,
                reveal_blind=reveal_blind,
            )
            for session in sessions
        ],
        "total": total,
        "limit": limit,
        "offset": offset,
    }


@router.post("/inventory-sessions", response_model=schemas.InventorySessionResponse)
def create_inventory_session(
    payload: schemas.InventorySessionCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    if payload.location_id:
        location = db.query(models.StockLocation).filter_by(id=payload.location_id, is_active=True).first()
        if not location:
            raise HTTPException(status_code=404, detail="Emplacement introuvable.")
    if not payload.name.strip():
        raise HTTPException(status_code=400, detail="Nom de campagne obligatoire.")
    inventory_type = str(payload.inventory_type or "full").strip().lower()
    if inventory_type not in {"full", "cycle"}:
        raise HTTPException(
            status_code=400,
            detail="Type d'inventaire invalide. Utilisez full ou cycle.",
        )
    assigned_usernames = _normalize_usernames(payload.assigned_usernames)
    _assert_known_inventory_counters(db, assigned_usernames)
    scheduled_for = _naive_utc(payload.scheduled_for)
    is_scheduled = bool(scheduled_for and scheduled_for > utcnow())
    # Gel de zone imposé à True par défaut côté serveur : le client peut
    # explicitement demander False, la garde anti-dérive 409 à la validation
    # reste alors le filet (cf. doc du schéma InventorySessionCreate).
    zone_locked = False if is_scheduled else (
        True if payload.zone_locked is None else bool(payload.zone_locked)
    )
    reference = f"INV-{int(time.time() * 1000)}"
    session = models.InventorySession(
        reference=reference,
        name=payload.name.strip(),
        location_id=payload.location_id,
        notes=payload.notes,
        status="scheduled" if is_scheduled else "draft",
        zone_locked=zone_locked,
        blind_counting=bool(payload.blind_counting),
        include_all_variants=bool(payload.include_all_variants),
        inventory_type=inventory_type,
        scheduled_for=scheduled_for,
        cycle_frequency_days=payload.cycle_frequency_days,
        assigned_usernames=assigned_usernames,
        approval_threshold_value=payload.approval_threshold_value,
        locked_at=utcnow() if zone_locked else None,
        unlocked_at=None if zone_locked else utcnow(),
        created_by=user.get("sub", "Admin"),
    )
    db.add(session)
    db.flush()
    if not is_scheduled:
        _prefill_inventory_lines(db, session, payload.include_all_variants)
    db.commit()
    created = _inventory_session_query(db).filter(
        models.InventorySession.id == session.id
    ).first()
    return _serialize_inventory_session(
        created,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.patch("/inventory-sessions/{session_id}", response_model=schemas.InventorySessionResponse)
def update_inventory_session(
    session_id: int,
    payload: schemas.InventorySessionUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status in {"validated", "cancelled"}:
        raise HTTPException(status_code=400, detail="Campagne clôturée.")
    changes = payload.model_dump(exclude_unset=True)
    if "assigned_usernames" in changes:
        assigned = _normalize_usernames(changes["assigned_usernames"])
        _assert_known_inventory_counters(db, assigned)
        session.assigned_usernames = assigned
    if "notes" in changes:
        session.notes = changes["notes"]
    if "approval_threshold_value" in changes:
        session.approval_threshold_value = changes["approval_threshold_value"]
        session.finance_approved_by = None
        session.finance_approved_at = None
        if session.status == "pending_approval":
            session.status = "counting"
    if "scheduled_for" in changes:
        if session.status != "scheduled":
            raise HTTPException(
                status_code=400,
                detail="La date ne peut être modifiée que pour une campagne planifiée.",
            )
        session.scheduled_for = _naive_utc(changes["scheduled_for"])
    db.commit()
    loaded = _inventory_session_query(db).filter(
        models.InventorySession.id == session.id
    ).first()
    return _serialize_inventory_session(
        loaded,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.post("/inventory-sessions/{session_id}/start", response_model=schemas.InventorySessionResponse)
def start_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status != "scheduled":
        raise HTTPException(status_code=409, detail="Cette campagne n'est pas planifiée.")
    session.status = "draft"
    session.zone_locked = True
    session.locked_at = utcnow()
    session.unlocked_at = None
    if not session.lines:
        _prefill_inventory_lines(db, session, bool(session.include_all_variants))
    db.commit()
    loaded = _inventory_session_query(db).filter(
        models.InventorySession.id == session.id
    ).first()
    return _serialize_inventory_session(
        loaded,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.get("/inventory-sessions/{session_id}", response_model=schemas.InventorySessionResponse)
def get_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    session = (
        _inventory_session_query(db)
        .filter(models.InventorySession.id == session_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    return _serialize_inventory_session(
        session,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.post("/inventory-sessions/{session_id}/lines", response_model=schemas.InventoryCountLineResponse)
def upsert_inventory_count_line(
    session_id: int,
    payload: schemas.InventoryCountLineUpsert,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.count")
    session = (
        db.query(models.InventorySession)
        .filter_by(id=session_id)
        .with_for_update()
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status not in ["draft", "counting"]:
        raise HTTPException(
            status_code=400,
            detail="La campagne n'est pas ouverte à la saisie.",
        )
    _assert_counter_assignment(session, user)
    if payload.client_operation_id:
        replayed = (
            db.query(models.InventoryCountLine)
            .filter(
                models.InventoryCountLine.last_client_operation_id
                == payload.client_operation_id
            )
            .first()
        )
        if replayed:
            if replayed.session_id != session_id:
                raise HTTPException(
                    status_code=409,
                    detail="Identifiant de synchronisation déjà utilisé.",
                )
            return _serialize_count_line(
                session,
                replayed,
                reveal_blind=_can_reveal_blind_inventory(db, user),
            )
    variant = db.query(models.ProductVariant).filter_by(id=payload.variant_id).first()
    if not variant:
        raise HTTPException(status_code=404, detail="Variante introuvable.")
    location = db.query(models.StockLocation).filter_by(id=payload.location_id, is_active=True).first()
    if not location:
        raise HTTPException(status_code=404, detail="Emplacement introuvable.")
    if location.usage != "internal":
        raise HTTPException(status_code=400, detail="Le comptage physique doit viser un emplacement interne.")
    if session.location_id and payload.location_id not in _zone_location_ids(db, session.location_id):
        raise HTTPException(status_code=400, detail="Cette campagne est limitée à un autre emplacement.")
    if payload.counted_quantity < 0:
        raise HTTPException(status_code=400, detail="La quantité comptée ne peut pas être négative.")

    expected = _get_quant_quantity(db, payload.variant_id, payload.location_id)
    counted = float(payload.counted_quantity)
    line = (
        db.query(models.InventoryCountLine)
        .filter_by(session_id=session_id, variant_id=payload.variant_id, location_id=payload.location_id)
        .with_for_update()
        .first()
    )
    if not line:
        line = models.InventoryCountLine(
            session_id=session_id,
            variant_id=payload.variant_id,
            location_id=payload.location_id,
            version=1,
        )
        db.add(line)
        expected = _get_quant_quantity(db, payload.variant_id, payload.location_id)
    else:
        if payload.expected_version is None and line.status != "pending":
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Version de ligne obligatoire. Rechargez la campagne.",
                    "current_version": int(line.version or 1),
                },
            )
        if (
            payload.expected_version is not None
            and int(payload.expected_version) != int(line.version or 1)
        ):
            raise HTTPException(
                status_code=409,
                detail={
                    "message": "Cette ligne a été modifiée par un autre compteur.",
                    "current_version": int(line.version or 1),
                    "counted_by": line.counted_by,
                    "counted_at": line.counted_at.isoformat() if line.counted_at else None,
                },
            )
        expected = float(line.expected_quantity or 0)
        line.version = int(line.version or 1) + 1
    line.expected_quantity = expected
    line.counted_quantity = counted
    line.variance_quantity = counted - expected
    line.status = _line_status_from_variance(line.variance_quantity)
    line.reason = payload.reason
    line.notes = payload.notes
    line.recount_requested_by = None
    line.recount_requested_at = None
    line.recount_notes = None
    line.counted_by = user.get("sub", "Admin")
    line.counted_at = utcnow()
    line.last_client_operation_id = payload.client_operation_id
    line.unit_cost_snapshot = float(variant.cost_price or 0)
    line.variance_value = float(line.variance_quantity or 0) * float(
        line.unit_cost_snapshot or 0
    )
    session.status = "counting"
    session.finance_approved_by = None
    session.finance_approved_at = None
    try:
        db.commit()
    except IntegrityError as exc:
        db.rollback()
        raise HTTPException(
            status_code=409,
            detail="Une saisie concurrente existe déjà pour cette référence et cet emplacement.",
        ) from exc
    db.refresh(line)
    return _serialize_count_line(
        session,
        line,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.post(
    "/inventory-sessions/{session_id}/lines/{line_id}/attachments",
    response_model=schemas.InventoryCountAttachmentResponse,
)
async def upload_inventory_count_attachment(
    session_id: int,
    line_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.count")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status not in {"draft", "counting", "pending_approval"}:
        raise HTTPException(status_code=400, detail="Campagne clôturée.")
    _assert_counter_assignment(session, user)
    line = (
        db.query(models.InventoryCountLine)
        .filter_by(id=line_id, session_id=session_id)
        .first()
    )
    if not line:
        raise HTTPException(status_code=404, detail="Ligne de comptage introuvable.")

    content_type = (file.content_type or "application/octet-stream").lower()
    allowed = (
        content_type.startswith("image/")
        or content_type in {"application/pdf", "text/plain"}
    )
    if not allowed:
        raise HTTPException(
            status_code=415,
            detail="Justificatif accepté: image, PDF ou fichier texte.",
        )
    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Le justificatif est vide.")
    if len(content) > 10 * 1024 * 1024:
        raise HTTPException(status_code=413, detail="Justificatif limité à 10 Mo.")

    original_name = Path(file.filename or "justificatif").name
    if content_type == "application/pdf":
        suffix = ".pdf"
    elif content_type == "text/plain":
        suffix = ".txt"
    else:
        image_suffixes = {
            ".avif",
            ".bmp",
            ".gif",
            ".heic",
            ".heif",
            ".jpeg",
            ".jpg",
            ".png",
            ".tif",
            ".tiff",
            ".webp",
        }
        original_suffix = Path(original_name).suffix.lower()
        suffix = original_suffix if original_suffix in image_suffixes else ".img"
    stored_filename = f"{uuid4().hex}{suffix}"
    target_dir = Path("uploads") / "inventory" / str(session_id) / str(line_id)
    target_dir.mkdir(parents=True, exist_ok=True)
    (target_dir / stored_filename).write_bytes(content)
    attachment = models.InventoryCountAttachment(
        line_id=line.id,
        filename=original_name,
        stored_filename=stored_filename,
        content_type=content_type,
        size_bytes=len(content),
        uploaded_by=_actor(user),
        created_at=utcnow(),
    )
    db.add(attachment)
    db.commit()
    db.refresh(attachment)
    response = schemas.InventoryCountAttachmentResponse.model_validate(attachment)
    response.url = (
        f"/uploads/inventory/{session_id}/{line_id}/{stored_filename}"
    )
    return response


@router.post("/inventory-sessions/{session_id}/lines/{line_id}/recount", response_model=schemas.InventoryCountLineResponse)
def request_inventory_line_recount(
    session_id: int,
    line_id: int,
    payload: schemas.InventoryRecountRequest,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status in ["validated", "cancelled"]:
        raise HTTPException(status_code=400, detail="Campagne clôturée.")
    line = (
        db.query(models.InventoryCountLine)
        .filter_by(id=line_id, session_id=session_id)
        .first()
    )
    if not line:
        raise HTTPException(status_code=404, detail="Ligne de comptage introuvable.")
    if abs(float(line.variance_quantity or 0)) <= 0.000001:
        raise HTTPException(status_code=400, detail="Cette ligne est déjà conforme.")
    line.status = "recount"
    line.recount_requested_by = user.get("sub", "Admin")
    line.recount_requested_at = utcnow()
    line.recount_notes = payload.notes
    session.status = "counting"
    session.finance_approved_by = None
    session.finance_approved_at = None
    db.commit()
    db.refresh(line)
    return _serialize_count_line(
        session,
        line,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.post(
    "/inventory-sessions/{session_id}/approve-value",
    response_model=schemas.InventorySessionResponse,
)
def approve_inventory_variance_value(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.approve_value")
    session = (
        _inventory_session_query(db)
        .filter(models.InventorySession.id == session_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status != "pending_approval":
        raise HTTPException(
            status_code=409,
            detail="Cette campagne n'attend pas d'approbation financière.",
        )
    _total_value, absolute_value = _inventory_value_summary(session)
    threshold = float(session.approval_threshold_value or 0)
    if threshold <= 0 or absolute_value <= threshold + 0.005:
        raise HTTPException(
            status_code=409,
            detail="Le seuil d'approbation n'est plus dépassé.",
        )
    session.finance_approved_by = _actor(user)
    session.finance_approved_at = utcnow()
    db.commit()
    loaded = _inventory_session_query(db).filter(
        models.InventorySession.id == session.id
    ).first()
    return _serialize_inventory_session(
        loaded,
        reveal_blind=_can_reveal_blind_inventory(db, user),
    )


@router.post("/inventory-sessions/{session_id}/validate", response_model=schemas.InventorySessionResponse)
def validate_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = (
        db.query(models.InventorySession)
        .options(joinedload(models.InventorySession.lines))
        .filter(models.InventorySession.id == session_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status == "validated":
        raise HTTPException(status_code=409, detail="Campagne déjà validée.")
    if session.status == "cancelled":
        raise HTTPException(status_code=400, detail="Campagne annulée.")
    if session.status == "scheduled":
        raise HTTPException(
            status_code=400,
            detail="Démarrez la campagne planifiée avant de la valider.",
        )
    if not session.lines:
        raise HTTPException(status_code=400, detail="Aucune ligne comptée à valider.")
    recount_lines = [line for line in session.lines if line.status == "recount"]
    if recount_lines:
        raise HTTPException(
            status_code=400,
            detail=f"{len(recount_lines)} ligne(s) sont en attente de recompte.",
        )
    pending_lines = [line for line in session.lines if line.status == "pending"]
    if pending_lines:
        raise HTTPException(
            status_code=400,
            detail=f"{len(pending_lines)} ligne(s) restent à compter avant validation.",
        )

    inventory_location = _get_or_create_inventory_location(db)
    author = user.get("sub", "Admin")
    for line in session.lines:
        current_qty = _get_quant_quantity(db, line.variant_id, line.location_id)
        if abs(current_qty - float(line.expected_quantity or 0)) > 0.000001:
            raise HTTPException(
                status_code=409,
                detail=(
                    "Stock modifié depuis le comptage. "
                    f"Relancez la ligne {line.variant_id} avant validation."
                ),
            )
        variance = float(line.variance_quantity or 0)
        if abs(variance) > 0.000001 and not (line.reason or "").strip():
            raise HTTPException(
                status_code=400,
                detail=(
                    "Motif obligatoire pour valider un écart d'inventaire. "
                    f"Ligne {line.id}: renseignez la raison du comptage."
                ),
            )

    _total_value, absolute_value = _inventory_value_summary(session)
    threshold = float(session.approval_threshold_value or 0)
    if (
        threshold > 0
        and absolute_value > threshold + 0.005
        and not session.finance_approved_at
    ):
        session.status = "pending_approval"
        db.commit()
        return get_inventory_session(session.id, db, user)

    # Verrou anti double validation concurrente : bascule atomique du statut.
    # Si une requête concurrente a déjà validé la campagne, rowcount vaut 0 et
    # toute la transaction (ajustements inclus) est annulée avec ce 409.
    updated_rows = db.execute(
        update(models.InventorySession)
        .where(
            models.InventorySession.id == session.id,
            models.InventorySession.status.in_(["draft", "counting", "pending_approval"]),
        )
        .values(status="validated")
    ).rowcount
    if not updated_rows:
        raise HTTPException(status_code=409, detail="Campagne déjà validée par une opération concurrente.")

    for line in session.lines:
        variance = float(line.variance_quantity or 0)
        if abs(variance) <= 0.000001:
            continue

        if variance > 0:
            source_id = inventory_location.id
            dest_id = line.location_id
            quantity = variance
        else:
            adjustment = abs(variance)
            source_id = line.location_id
            dest_id = inventory_location.id
            quantity = adjustment

        try:
            result = InventoryService.move_stock(
                db,
                variant_id=line.variant_id,
                source_location_id=source_id,
                dest_location_id=dest_id,
                quantity=quantity,
                reference=f"INV/{session.reference}/{line.id}",
                notes=f"Ajustement inventaire physique {session.reference}. Motif: {line.reason or 'Non renseigné'}",
                author=author,
                source_screen="stock.physical_inventory",
                document_type="inventory_session",
                document_reference=session.reference,
                business_reason=line.reason,
                enforce_zone_lock=False,
            )
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=f"Ajustement impossible: {exc}") from exc
        line.adjustment_move_id = result.move.id
        line.status = "validated"

    for line in session.lines:
        line.status = "validated"

    session.status = "validated"
    session.validated_by = author
    session.validated_at = utcnow()
    session.zone_locked = False
    session.unlocked_at = utcnow()
    if (
        session.inventory_type == "cycle"
        and session.cycle_frequency_days
        and session.cycle_frequency_days > 0
    ):
        db.add(models.InventorySession(
            reference=f"INV-{int(time.time() * 1000)}-{uuid4().hex[:4].upper()}",
            name=f"{session.name} · cycle suivant",
            status="scheduled",
            location_id=session.location_id,
            notes=session.notes,
            zone_locked=False,
            blind_counting=session.blind_counting,
            include_all_variants=session.include_all_variants,
            inventory_type="cycle",
            scheduled_for=utcnow() + timedelta(days=session.cycle_frequency_days),
            cycle_frequency_days=session.cycle_frequency_days,
            assigned_usernames=list(session.assigned_usernames or []),
            approval_threshold_value=session.approval_threshold_value,
            locked_at=None,
            unlocked_at=utcnow(),
            created_by=author,
        ))
    db.commit()
    db.refresh(session)
    return get_inventory_session(session.id, db, user)


@router.post("/inventory-sessions/{session_id}/cancel", response_model=schemas.InventorySessionResponse)
def cancel_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status == "validated":
        raise HTTPException(status_code=400, detail="Campagne déjà validée.")
    session.status = "cancelled"
    session.zone_locked = False
    session.unlocked_at = utcnow()
    db.commit()
    db.refresh(session)
    return session


@router.post(
    "/inventory-sessions/{session_id}/archive",
    response_model=schemas.InventorySessionResponse,
)
def archive_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    if session.status not in {"validated", "cancelled"}:
        raise HTTPException(
            status_code=409,
            detail="Seules les campagnes clôturées peuvent être archivées.",
        )
    session.archived_by = _actor(user)
    session.archived_at = utcnow()
    db.commit()
    return get_inventory_session(session.id, db, user)


@router.post(
    "/inventory-sessions/{session_id}/restore",
    response_model=schemas.InventorySessionResponse,
)
def restore_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "inventory.validate")
    session = db.query(models.InventorySession).filter_by(id=session_id).first()
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")
    session.archived_by = None
    session.archived_at = None
    db.commit()
    return get_inventory_session(session.id, db, user)


@router.get("/inventory-sessions/{session_id}/export")
def export_inventory_session(
    session_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    # L'export embarque espéré/compté/écart en clair : en comptage aveugle, un
    # simple compteur ne doit pas pouvoir s'exporter les espérés. On exige la
    # même permission que la validation de campagne (miroir du masquage de
    # ``_serialize_inventory_session``).
    _require_permission(db, user, "inventory.validate")
    session = (
        _inventory_session_query(db)
        .filter(models.InventorySession.id == session_id)
        .first()
    )
    if not session:
        raise HTTPException(status_code=404, detail="Campagne d'inventaire introuvable.")

    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.title = "Inventaire physique"
    sheet.append(["Campagne", session.reference, session.name])
    sheet.append(["Statut", session.status, "Zone gelée" if session.zone_locked else "Zone libérée"])
    sheet.append(["Emplacement", session.location.name if session.location else "Tous emplacements internes"])
    sheet.append(["Compteurs affectés", ", ".join(session.assigned_usernames or []) or "Tous"])
    total_value, absolute_value = _inventory_value_summary(session)
    sheet.append(["Écart valorisé net", float(total_value), "Valeur absolue", float(absolute_value)])
    sheet.append([
        "Seuil d'approbation",
        float(session.approval_threshold_value or 0),
        "Approuvé par",
        session.finance_approved_by or "",
    ])
    sheet.append([])
    headers = [
        "Référence",
        "Produit",
        "Emplacement",
        "Système",
        "Compté",
        "Écart",
        "Coût unitaire",
        "Écart valorisé",
        "Statut ligne",
        "Motif",
        "Compté par",
        "Recompte demandé par",
        "Note recompte",
        "Justificatifs",
        "Mouvement ajustement",
    ]
    sheet.append(headers)
    header_row = sheet.max_row
    for cell in sheet[header_row]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1E293B")

    for line in session.lines:
        variant = line.variant
        product = variant.product if variant else None
        sheet.append([
            variant.reference if variant else f"Variante #{line.variant_id}",
            product.name if product else "",
            line.location.name if line.location else f"Lieu #{line.location_id}",
            float(line.expected_quantity or 0),
            float(line.counted_quantity or 0),
            float(line.variance_quantity or 0),
            float(line.unit_cost_snapshot or 0),
            _line_variance_value(line),
            line.status,
            line.reason or "",
            line.counted_by or "",
            line.recount_requested_by or "",
            line.recount_notes or "",
            ", ".join(attachment.filename for attachment in line.attachments),
            line.adjustment_move_id or "",
        ])

    for column in sheet.columns:
        max_length = max(len(str(cell.value or "")) for cell in column)
        sheet.column_dimensions[column[0].column_letter].width = min(max(max_length + 2, 12), 44)

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    filename = f"{session.reference}-rapport-inventaire.xlsx"
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename}"'},
    )

@router.get("/products", response_model=List[schemas.ProductResponse])
def get_products(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    products = db.query(models.Product).options(joinedload(models.Product.variants)).all()
    blind_locations = _active_blind_location_ids(db, user)
    responses = []
    for product in products:
        for variant in product.variants:
            annotate_variant_availability(db, variant)
        response = schemas.ProductResponse.model_validate(product)
        if blind_locations:
            for variant in response.variants:
                variant.quantity_in_stock = None
                variant.reserved_quantity = None
                variant.available_quantity = None
        responses.append(response)
    return responses

@router.post("/products", response_model=schemas.ProductResponse)
def create_product(product_data: schemas.ProductCreate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")

    reference_base = product_data.reference_base.strip().upper()
    existing = db.query(models.Product).filter(
        func.upper(models.Product.reference_base) == reference_base
    ).first()
    if existing:
        raise HTTPException(409, f"La référence famille {reference_base} existe déjà.")

    product_values = product_data.model_dump(exclude={'variants'})
    product_values["reference_base"] = reference_base
    product_values["name"] = product_data.name.strip()
    product_values["supplier"] = _clean(product_data.supplier)
    if not product_values.get("category"):
        product_values["category"] = "SERVICE" if product_data.product_type == "service" else product_data.material_type
    product_values["catalog_status"] = (product_data.catalog_status or "DRAFT").upper()
    if product_values["catalog_status"] not in CATALOG_STATUSES:
        raise HTTPException(400, "Statut catalogue inconnu.")

    new_product = models.Product(**product_values)
    db.add(new_product)
    db.flush()

    for v_data in product_data.variants:
        variant_values = v_data.model_dump()
        variant_values["reference"] = v_data.reference.strip().upper()
        variant_values["barcode"] = _clean(v_data.barcode)
        variant_values["supplier_reference"] = _clean(v_data.supplier_reference)
        _ensure_variant_uniqueness(
            db,
            reference=variant_values["reference"],
            barcode=variant_values["barcode"],
            supplier_reference=variant_values["supplier_reference"],
            supplier=new_product.supplier,
        )
        new_variant = models.ProductVariant(product_id=new_product.id, **variant_values)
        db.add(new_variant)
        db.flush()

    _record_product_audit(
        db,
        product_id=new_product.id,
        user=user,
        action="PRODUCT_CREATED",
        changes={"catalog_status": {"before": None, "after": new_product.catalog_status}},
    )
    db.commit()
    db.refresh(new_product)
    return new_product

@router.put("/products/{product_id}", response_model=schemas.ProductResponse)
def update_product(product_id: int, product_data: schemas.ProductBase, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")
        
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product: raise HTTPException(404, "Product not found")
    product_values = product_data.model_dump()
    requested_status = (product_values.pop("catalog_status", None) or product.catalog_status).upper()
    if requested_status != product.catalog_status:
        raise HTTPException(400, "Utilisez l'action de changement de statut catalogue.")
    if "category" not in product_data.model_fields_set:
        product_values.pop("category", None)
    elif not product_values.get("category"):
        product_values["category"] = "SERVICE" if product_data.product_type == "service" else product_data.material_type
    product_values["reference_base"] = product_data.reference_base.strip().upper()
    product_values["name"] = product_data.name.strip()
    product_values["supplier"] = _clean(product_data.supplier)

    duplicate = db.query(models.Product).filter(
        func.upper(models.Product.reference_base) == product_values["reference_base"],
        models.Product.id != product_id,
    ).first()
    if duplicate:
        raise HTTPException(409, f"La référence famille {product_values['reference_base']} existe déjà.")

    changes = _changes(product, product_values)
    for key, value in product_values.items():
        setattr(product, key, value)
    if changes:
        _record_product_audit(
            db,
            product_id=product.id,
            user=user,
            action="PRODUCT_UPDATED",
            changes=changes,
        )
    db.commit()
    db.refresh(product)
    return product


@router.post("/products/{product_id}/status", response_model=schemas.ProductResponse)
def transition_product_status(
    product_id: int,
    payload: schemas.ProductStatusUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "catalog.qualify")
    product = (
        db.query(models.Product)
        .options(joinedload(models.Product.variants))
        .filter(models.Product.id == product_id)
        .first()
    )
    if not product:
        raise HTTPException(404, "Product not found")

    current_status = (product.catalog_status or "DRAFT").upper()
    next_status = payload.status.upper()
    if next_status not in CATALOG_STATUSES:
        raise HTTPException(400, "Statut catalogue inconnu.")
    if next_status == current_status:
        return product
    if next_status not in CATALOG_TRANSITIONS.get(current_status, set()):
        raise HTTPException(409, f"Transition {current_status} vers {next_status} interdite.")
    if next_status == "ACTIVE":
        issues = _activation_issues(product)
        if issues:
            raise HTTPException(
                422,
                "Activation impossible. À compléter : " + ", ".join(issues) + ".",
            )
    if next_status in {"BLOCKED", "ARCHIVED"} and not _clean(payload.reason):
        raise HTTPException(422, "Une raison est obligatoire pour bloquer ou archiver un article.")

    product.catalog_status = next_status
    _record_product_audit(
        db,
        product_id=product.id,
        user=user,
        action="STATUS_CHANGED",
        changes={"catalog_status": {"before": current_status, "after": next_status}},
        reason=payload.reason,
    )
    db.commit()
    db.refresh(product)
    return product


@router.get("/products/{product_id}/history", response_model=List[schemas.ProductAuditLogResponse])
def get_product_history(
    product_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    product = db.query(models.Product.id).filter(models.Product.id == product_id).first()
    if not product:
        raise HTTPException(404, "Product not found")
    return (
        db.query(models.ProductAuditLog)
        .filter(models.ProductAuditLog.product_id == product_id)
        .order_by(models.ProductAuditLog.created_at.desc(), models.ProductAuditLog.id.desc())
        .limit(100)
        .all()
    )

import os
import uuid
import shutil
from ..core import uploads

@router.post("/products/upload_image")
async def upload_product_image(file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")

    filepath = await uploads.save_upload_file(file, os.path.join("uploads", "products"))
    return {"image_url": f"/uploads/products/{os.path.basename(filepath)}"}


@router.put("/variants/{variant_id}", response_model=schemas.ProductVariantResponse)
def update_variant(variant_id: int, variant_data: schemas.ProductVariantBase, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")
        
    variant = db.query(models.ProductVariant).filter(models.ProductVariant.id == variant_id).first()
    if not variant: raise HTTPException(404, "Variant not found")
    variant_values = variant_data.model_dump()
    variant_values["reference"] = variant_data.reference.strip().upper()
    variant_values["barcode"] = _clean(variant_data.barcode)
    variant_values["supplier_reference"] = _clean(variant_data.supplier_reference)
    _ensure_variant_uniqueness(
        db,
        reference=variant_values["reference"],
        barcode=variant_values["barcode"],
        supplier_reference=variant_values["supplier_reference"],
        supplier=variant.product.supplier if variant.product else None,
        exclude_variant_id=variant.id,
    )
    variant_values.pop("quantity_in_stock", None)
    changes = _changes(variant, variant_values)
    for key, value in variant_values.items():
        setattr(variant, key, value)
    if changes:
        _record_product_audit(
            db,
            product_id=variant.product_id,
            variant_id=variant.id,
            user=user,
            action="VARIANT_UPDATED",
            changes=changes,
        )
    db.commit()
    db.refresh(variant)
    annotate_variant_availability(db, variant)
    return variant

@router.post("/products/{product_id}/variants", response_model=schemas.ProductVariantResponse)
def add_variant(product_id: int, variant_data: schemas.ProductVariantCreate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")
    product = db.query(models.Product).filter(models.Product.id == product_id).first()
    if not product: raise HTTPException(404, "Product not found")
    variant_values = variant_data.model_dump()
    variant_values["reference"] = variant_data.reference.strip().upper()
    variant_values["barcode"] = _clean(variant_data.barcode)
    variant_values["supplier_reference"] = _clean(variant_data.supplier_reference)
    _ensure_variant_uniqueness(
        db,
        reference=variant_values["reference"],
        barcode=variant_values["barcode"],
        supplier_reference=variant_values["supplier_reference"],
        supplier=product.supplier,
    )
    new_variant = models.ProductVariant(product_id=product.id, **variant_values)
    db.add(new_variant)
    db.flush()
    _record_product_audit(
        db,
        product_id=product.id,
        variant_id=new_variant.id,
        user=user,
        action="VARIANT_CREATED",
        changes={"reference": {"before": None, "after": new_variant.reference}},
    )
    db.commit()
    db.refresh(new_variant)
    annotate_variant_availability(db, new_variant)
    return new_variant

from fastapi import BackgroundTasks

# ODOO ENGINE: Stock Moves
@router.post("/transaction") # Kept same endpoint name for UI compat momentarily, but treats it as an Odoo Move
def create_transaction(tx: schemas.StockMoveCreate, background_tasks: BackgroundTasks, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    from ..core.events import EventBus
        
    variant = db.query(models.ProductVariant).filter(models.ProductVariant.id == tx.variant_id).first()
    if not variant: raise HTTPException(404, "Variant not found")

    qty = abs(tx.quantity)
    src_loc = db.query(models.StockLocation).filter_by(id=tx.location_id).first() if tx.location_id else None
    dest_loc = db.query(models.StockLocation).filter_by(id=tx.location_dest_id).first() if tx.location_dest_id else None
    is_manual_inventory_adjustment = bool(
        (src_loc and src_loc.usage == "inventory")
        or (dest_loc and dest_loc.usage == "inventory")
    )
    if is_manual_inventory_adjustment:
        required_permission = "stock.adjust"
    elif src_loc and dest_loc and src_loc.usage == "internal" and dest_loc.usage == "internal":
        required_permission = "stock.transfer"
    elif dest_loc and dest_loc.usage == "internal":
        required_permission = "stock.receive"
    else:
        required_permission = "stock.adjust"
    _require_permission(db, user, required_permission)
    if is_manual_inventory_adjustment and not (tx.reason or "").strip():
        raise HTTPException(
            status_code=400,
            detail="Motif obligatoire pour un ajustement manuel d'inventaire.",
        )
    try:
        result = InventoryService.move_stock(
            db,
            variant_id=tx.variant_id,
            source_location_id=tx.location_id,
            dest_location_id=tx.location_dest_id,
            quantity=qty,
            reference=f"WH/MOVE-{int(time.time()*1000)}",
            notes=tx.notes,
            author=user.get("sub", "Admin"),
            source_screen=tx.source_screen or "stock.manual_transaction",
            document_type=tx.document_type or ("manual_inventory_adjustment" if is_manual_inventory_adjustment else "manual_stock_move"),
            document_reference=tx.document_reference,
            business_reason=tx.reason or tx.notes,
        )
    except ValueError as exc:
        detail = str(exc)
        status_code = 423 if "Zone gelée" in detail else 400
        raise HTTPException(status_code=status_code, detail=detail) from exc

    if (
        src_loc
        and src_loc.usage == 'internal'
        and result.previous_source_quantity is not None
        and result.new_source_quantity is not None
        and result.previous_source_quantity > variant.min_threshold
        and result.new_source_quantity <= variant.min_threshold
    ):
        EventBus.on_stock_alert(variant.reference, result.new_source_quantity, background_tasks)
    
    # --- LOG CHATTER (AUDIT) ---
    src_name = "Externe"
    if tx.location_id:
        src_name = src_loc.name if src_loc else "Inconnu"
        
    dest_name = "Externe"
    if tx.location_dest_id:
        dest_name = dest_loc.name if dest_loc else "Inconnu"

    msg = f"Mouvement de {qty} unité(s): {src_name} ➔ {dest_name} (Mvmt: {result.move.reference})"
    
    # Enrichment for Suppliers / Clients traceability
    variant_db = db.query(models.ProductVariant).filter(models.ProductVariant.id == tx.variant_id).first()
    supplier_name = variant_db.product.supplier if variant_db and variant_db.product and variant_db.product.supplier else "Fournisseur Inconnu"
    
    if tx.location_id and src_loc and src_loc.usage == 'supplier':
        msg += f"\nRéception depuis : {supplier_name}"
    elif tx.location_dest_id and dest_loc and dest_loc.usage == 'customer':
        # Default customer text for manual delivery
        msg += f"\nLivraison vers : Client"

    if tx.notes:
        msg += f"\nNote: {tx.notes}"

    audit_log = models.ChatterMessage(
        model_name="variant",
        record_id=tx.variant_id,
        body=msg,
        author=user.get("sub", "Admin"),
        is_system_log=True
    )
    db.add(audit_log)
    # ---------------------------

    db.commit()
    return {"status": "success"}

@router.get("/transactions")
def get_recent_transactions(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    moves = (
        db.query(models.StockMove)
        .options(
            joinedload(models.StockMove.variant).joinedload(models.ProductVariant.product),
            joinedload(models.StockMove.source_location),
            joinedload(models.StockMove.dest_location),
        )
        .order_by(models.StockMove.date.desc())
        .limit(100)
        .all()
    )
    result = []
    for m in moves:
        item_name = f"{m.variant.product.name} ({m.variant.color or 'Std'})" if m.variant and m.variant.product else "Inconnu"
        
        # Resolve names for display
        src_name = m.source_location.name if m.source_location else "Fournisseur / Externe"
        dest_name = m.dest_location.name if m.dest_location else "Client / Perte"
        is_workshop_debit = bool(
            (m.reference or "").startswith("DEBIT-ATELIER")
            or "Débit atelier réel" in (m.notes or "")
            or "Consommation réservation" in (m.notes or "")
        )
        
        result.append({
            "id": m.id,
            "reference": m.reference,
            "variant_id": m.variant_id,
            "item_name": item_name,
            "quantity_change": m.quantity,
            "transaction_type": "Débit atelier réel" if is_workshop_debit else f"{src_name} ➔ {dest_name}",
            "movement_kind": "workshop_debit" if is_workshop_debit else "stock_move",
            # Ids + noms des emplacements source/destination (None si virtuel :
            # réception fournisseur, livraison client, perte). Permet au
            # frontend de filtrer les mouvements d'une fiche emplacement.
            "location_id": m.location_id,
            "location_dest_id": m.location_dest_id,
            "location_from_name": m.source_location.name if m.source_location else None,
            "location_to_name": m.dest_location.name if m.dest_location else None,
            "created_at": m.date,
            "author": m.author or "Admin",
            "notes": m.notes,
            "source_screen": m.source_screen,
            "document_type": m.document_type,
            "document_reference": m.document_reference,
            "business_reason": m.business_reason,
        })
    return result


@router.get("/transactions/export")
def export_stock_audit_xlsx(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "stock.adjust")

    moves = (
        db.query(models.StockMove)
        .options(
            joinedload(models.StockMove.variant).joinedload(models.ProductVariant.product),
            joinedload(models.StockMove.source_location),
            joinedload(models.StockMove.dest_location),
        )
        .order_by(models.StockMove.date.desc())
        .limit(5000)
        .all()
    )
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Audit stock"
    headers = [
        "Date",
        "Reference mouvement",
        "Source ecran",
        "Type document",
        "Reference document",
        "Motif",
        "Auteur",
        "Article",
        "Reference variante",
        "Source",
        "Destination",
        "Quantite",
        "Notes",
    ]
    ws.append(headers)
    for cell in ws[1]:
        cell.font = Font(bold=True)
        cell.fill = PatternFill("solid", fgColor="E2E8F0")

    for move in moves:
        product_name = move.variant.product.name if move.variant and move.variant.product else "Inconnu"
        variant_ref = move.variant.reference if move.variant else ""
        ws.append(
            [
                move.date.strftime("%Y-%m-%d %H:%M:%S") if move.date else "",
                move.reference or "",
                move.source_screen or "",
                move.document_type or "",
                move.document_reference or "",
                move.business_reason or "",
                move.author or "",
                product_name,
                variant_ref,
                move.source_location.name if move.source_location else "Externe",
                move.dest_location.name if move.dest_location else "Externe",
                float(move.quantity or 0),
                move.notes or "",
            ]
        )

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = StreamingResponse(
        iter([stream.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    response.headers["Content-Disposition"] = "attachment; filename=stock-audit.xlsx"
    return response

@router.post("/workshop-debits/preview", response_model=schemas.WorkshopDebitPreviewResponse)
async def preview_workshop_debits(
    files: List[UploadFile] = File(...),
    source_location: str = Form("WH/Stock"),
    sale_order_id: Optional[int] = Form(None),
    production_order_id: Optional[int] = Form(None),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    records, issues, _source_names = await _parse_workshop_uploads(files)
    return build_preview_payload(
        db,
        records,
        issues,
        source_location,
        sale_order_id=sale_order_id,
        production_order_id=production_order_id,
    )

@router.post("/workshop-debits/draft-products")
async def create_draft_products_from_workshop_debits(
    files: List[UploadFile] = File(...),
    source_location: str = Form("WH/Stock"),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "catalog.qualify")

    records, issues, _source_names = await _parse_workshop_uploads(files)
    if any(issue.severity == "error" for issue in issues):
        raise HTTPException(status_code=400, detail="Fichier de débit non exploitable.")

    preview = build_preview_payload(db, records, issues, source_location)
    unknown_lines = [line for line in preview["stock_matches"] if line["status"] == "not_found"]
    location = db.query(models.StockLocation).filter_by(name=source_location, usage="internal").first()
    if not location:
        location = models.StockLocation(name=source_location, usage="internal", is_active=True)
        db.add(location)
        db.flush()

    created = 0
    skipped = 0
    refs = []
    for line in unknown_lines:
        supplier = (line["supplier"] or "FOURNISSEUR").strip().upper()
        reference = (line["reference"] or "").strip()
        if not reference:
            skipped += 1
            continue
        variant_ref = f"{supplier}:{reference}"
        existing = db.query(models.ProductVariant).filter(models.ProductVariant.reference == variant_ref).first()
        if existing:
            skipped += 1
            continue

        unit = line["unit"] or "unité"
        material_type = "ALU" if supplier in {"SEPALUMIC", "CORTIZO", "TECHNAL/HYDRO", "TECHNAL", "HYDRO"} else "UNKNOWN"
        product = models.Product(
            reference_base=variant_ref,
            name=f"[BROUILLON] {supplier} {reference} - à compléter",
            category="PROFIL" if unit == "barre" else "ACCESSOIRE",
            material_type=material_type,
            unit=unit,
            supplier=supplier,
            product_type="stockable",
            available_in_pos=False,
            compatible_series="Créé depuis prévisualisation débit atelier; stock réel à renseigner.",
            catalog_status="DRAFT",
        )
        db.add(product)
        db.flush()
        variant = models.ProductVariant(
            product_id=product.id,
            reference=variant_ref,
            supplier_reference=reference,
            quantity_in_stock=0,
            min_threshold=0,
            location=source_location,
        )
        db.add(variant)
        _record_product_audit(
            db,
            product_id=product.id,
            user=user,
            action="PRODUCT_CREATED_FROM_WORKSHOP",
            changes={"catalog_status": {"before": None, "after": "DRAFT"}},
            reason="Référence inconnue détectée dans un fichier de débit atelier.",
        )
        db.flush()
        db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=0))
        created += 1
        refs.append(variant_ref)

    db.commit()
    return {
        "created": created,
        "skipped": skipped,
        "references": refs[:50],
        "message": f"{created} brouillon(s) catalogue créé(s), {skipped} ignoré(s).",
    }

@router.get("/workshop-debits/contexts")
def list_workshop_debit_contexts(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    sales = (
        db.query(models.SaleOrder)
        .filter(models.SaleOrder.status.in_(["SENT", "VALIDATED", "IN_DESIGN", "READY_FOR_PROD", "IN_PRODUCTION"]))
        .order_by(models.SaleOrder.updated_at.desc())
        .limit(100)
        .all()
    )
    active_plans = (
        db.query(models.Planning)
        .options(joinedload(models.Planning.order))
        .filter(
            models.Planning.status.in_([
                models.PlanningStatus.PENDING,
                models.PlanningStatus.IN_PROGRESS,
                models.PlanningStatus.PAUSED,
                models.PlanningStatus.ISSUE,
            ])
        )
        .order_by(models.Planning.created_at.desc())
        .limit(100)
        .all()
    )

    production_orders = []
    seen_order_ids = set()
    for plan in active_plans:
        if not plan.order or plan.order_id in seen_order_ids:
            continue
        seen_order_ids.add(plan.order_id)
        material = plan.order.material.value if hasattr(plan.order.material, "value") else plan.order.material
        production_orders.append(
            {
                "type": "production_order",
                "id": plan.order.id,
                "reference": plan.order.reference,
                "client_name": plan.order.client_name,
                "status": plan.status.value if hasattr(plan.status, "value") else plan.status,
                "material": material,
                "station": plan.station,
                "label": f"{plan.order.reference} - {plan.order.client_name or 'Atelier'} ({material})",
            }
        )

    return {
        "sales": [
            {
                "type": "sale_order",
                "id": sale.id,
                "reference": sale.reference,
                "client_name": sale.client_name,
                "status": sale.status,
                "is_reservable": sale.status in ["VALIDATED", "IN_DESIGN", "READY_FOR_PROD", "IN_PRODUCTION"],
                "label": f"{sale.reference} - {sale.client_name}",
            }
            for sale in sales
        ],
        "production_orders": production_orders,
    }

@router.post("/workshop-debits/reservations", response_model=schemas.StockReservationResponse)
async def reserve_workshop_debits(
    files: List[UploadFile] = File(...),
    source_location: str = Form("WH/Stock"),
    sale_order_id: Optional[int] = Form(None),
    production_order_id: Optional[int] = Form(None),
    order_reference: Optional[str] = Form(None),
    notes: Optional[str] = Form(None),
    allow_missing: bool = Form(False),
    allow_shortage: bool = Form(False),
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "workshop.reserve_stock")
    if (allow_missing or allow_shortage) and user.get("role") != "ADMIN":
        raise HTTPException(status_code=403, detail="Seul un administrateur peut forcer une réservation incomplète.")
    records, issues, source_names = await _parse_workshop_uploads(files)
    blocking_errors = [issue for issue in issues if issue.severity == "error"]
    if blocking_errors:
        raise HTTPException(status_code=400, detail="Fichier de débit non exploitable.")
    try:
        reservation = create_reservation(
            db,
            records,
            source_label=", ".join(source_names),
            created_by=user.get("sub", "Admin"),
            source_location=source_location,
            order_reference=order_reference,
            sale_order_id=sale_order_id,
            production_order_id=production_order_id,
            notes=notes,
            allow_missing=allow_missing,
            allow_shortage=allow_shortage,
        )
        if sale_order_id:
            # Aligné sur prepare-workshop/reserve : la réservation atelier fait
            # passer le devis en READY_FOR_PROD. Jamais de rétrogradation.
            sale = db.query(models.SaleOrder).filter(models.SaleOrder.id == sale_order_id).first()
            if sale and sale.status in ["VALIDATED", "IN_DESIGN"]:
                sale.status = "READY_FOR_PROD"
        db.commit()
        db.refresh(reservation)
        return reservation
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc))

@router.get("/workshop-debits/reservations", response_model=List[schemas.StockReservationResponse])
def list_workshop_reservations(
    status: Optional[str] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    query = db.query(models.StockReservation).options(
        joinedload(models.StockReservation.lines).joinedload(models.StockReservationLine.variant)
    ).order_by(models.StockReservation.created_at.desc())
    if status:
        query = query.filter(models.StockReservation.status == status)
    return query.limit(50).all()


@router.get("/workshop-preparations", response_model=List[schemas.WorkshopPreparationResponse])
def list_workshop_preparations(
    status: Optional[str] = None,
    reservation_id: Optional[int] = None,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    query = (
        db.query(models.WorkshopPreparation)
        .options(
            joinedload(models.WorkshopPreparation.lines)
            .joinedload(models.WorkshopPreparationLine.variant)
            .joinedload(models.ProductVariant.product),
            joinedload(models.WorkshopPreparation.source_location),
            joinedload(models.WorkshopPreparation.destination_location),
        )
        .order_by(models.WorkshopPreparation.created_at.desc())
    )
    if status:
        query = query.filter(models.WorkshopPreparation.status == status)
    if reservation_id:
        query = query.filter(models.WorkshopPreparation.reservation_id == reservation_id)
    return query.limit(100).all()


@router.post("/workshop-preparations", response_model=schemas.WorkshopPreparationResponse)
def create_workshop_preparation(
    payload: schemas.WorkshopPreparationCreate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "stock.transfer")
    reservation = (
        db.query(models.StockReservation)
        .options(joinedload(models.StockReservation.lines))
        .filter(models.StockReservation.id == payload.reservation_id)
        .first()
    )
    if not reservation:
        raise HTTPException(status_code=404, detail="Réservation introuvable.")
    try:
        preparation = create_preparation(
            db,
            reservation=reservation,
            destination_location_id=payload.destination_location_id,
            notes=payload.notes,
            author=_actor(user),
        )
        db.commit()
        return load_preparation(db, preparation.id) or preparation
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.patch(
    "/workshop-preparations/{preparation_id}/lines/{line_id}",
    response_model=schemas.WorkshopPreparationResponse,
)
def update_workshop_preparation_line(
    preparation_id: int,
    line_id: int,
    payload: schemas.WorkshopPreparationLineUpdate,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "stock.transfer")
    preparation = load_preparation(db, preparation_id, for_update=True)
    if not preparation:
        raise HTTPException(status_code=404, detail="Bon de préparation introuvable.")
    try:
        update_prepared_quantity(db, preparation, line_id, payload.prepared_quantity)
        db.commit()
        return load_preparation(db, preparation.id) or preparation
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/workshop-preparations/{preparation_id}/handover")
def hand_over_workshop_preparation(
    preparation_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "stock.transfer")
    preparation = load_preparation(db, preparation_id, for_update=True)
    if not preparation:
        raise HTTPException(status_code=404, detail="Bon de préparation introuvable.")
    try:
        stats = hand_over_preparation(db, preparation, author=_actor(user))
        db.commit()
        return {"status": "success", "reference": preparation.reference, **stats}
    except InsufficientStockAtConsumptionError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail=str(exc)) from exc
    except ValueError as exc:
        db.rollback()
        status_code = 423 if "Zone gelée" in str(exc) else 400
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc


@router.post("/workshop-preparations/{preparation_id}/return")
def return_workshop_preparation(
    preparation_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "stock.transfer")
    preparation = load_preparation(db, preparation_id, for_update=True)
    if not preparation:
        raise HTTPException(status_code=404, detail="Bon de préparation introuvable.")
    try:
        stats = return_preparation(db, preparation, author=_actor(user))
        db.commit()
        return {"status": "success", "reference": preparation.reference, **stats}
    except ValueError as exc:
        db.rollback()
        status_code = 423 if "Zone gelée" in str(exc) else 400
        raise HTTPException(status_code=status_code, detail=str(exc)) from exc


@router.post("/workshop-preparations/{preparation_id}/cancel")
def cancel_workshop_preparation(
    preparation_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "stock.transfer")
    preparation = load_preparation(db, preparation_id, for_update=True)
    if not preparation:
        raise HTTPException(status_code=404, detail="Bon de préparation introuvable.")
    try:
        cancel_preparation(db, preparation)
        db.commit()
        return {"status": "success", "reference": preparation.reference}
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@router.post("/workshop-debits/reservations/{reservation_id}/consume")
def consume_workshop_reservation(
    reservation_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "workshop.consume_stock")
    reservation = (
        db.query(models.StockReservation)
        .options(joinedload(models.StockReservation.lines).joinedload(models.StockReservationLine.variant))
        .filter(models.StockReservation.id == reservation_id)
        .first()
    )
    if not reservation:
        raise HTTPException(status_code=404, detail="Réservation introuvable.")
    try:
        stats = consume_reservation(db, reservation, author=user.get("sub", "Admin"))
        db.commit()
        return {"status": "success", **stats}
    except InsufficientStockAtConsumptionError as exc:
        db.rollback()
        raise HTTPException(status_code=409, detail=str(exc))
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc))

@router.post("/workshop-debits/reservations/{reservation_id}/cancel")
def cancel_workshop_reservation(
    reservation_id: int,
    db: Session = Depends(get_db),
    user: dict = Depends(get_current_user),
):
    _require_permission(db, user, "workshop.reserve_stock")
    reservation = (
        db.query(models.StockReservation)
        .options(joinedload(models.StockReservation.lines))
        .filter(models.StockReservation.id == reservation_id)
        .first()
    )
    if not reservation:
        raise HTTPException(status_code=404, detail="Réservation introuvable.")
    try:
        stats = cancel_reservation(db, reservation)
        db.commit()
        return {"status": "success", **stats}
    except ValueError as exc:
        db.rollback()
        raise HTTPException(status_code=400, detail=str(exc)) from exc

@router.get("/locations", response_model=List[schemas.StockLocationResponse])
def get_locations(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    # On renvoie uniquement les emplacements actifs pour cacher les archivés
    return db.query(models.StockLocation).filter(models.StockLocation.is_active == True).all()

@router.post("/locations", response_model=schemas.StockLocationResponse)
def create_location(loc: schemas.StockLocationCreate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "stock.locations.manage")

    name = (loc.name or "").strip()
    if not name:
        raise HTTPException(status_code=400, detail="Nom d'emplacement obligatoire.")

    usage = loc.usage or "internal"
    parent_id = loc.parent_id
    if parent_id:
        parent = db.query(models.StockLocation).filter_by(id=parent_id, is_active=True).first()
        if not parent:
            raise HTTPException(status_code=400, detail="Zone parente introuvable ou archivée.")
        usage = parent.usage

    existing = (
        db.query(models.StockLocation)
        .filter(models.StockLocation.name == name, models.StockLocation.parent_id == parent_id)
        .first()
    )
    if existing:
        raise HTTPException(status_code=400, detail="Un emplacement porte déjà ce nom.")

    db_loc = models.StockLocation(
        name=name,
        usage=usage,
        parent_id=parent_id,
        is_active=loc.is_active,
    )
    db.add(db_loc)
    db.commit()
    db.refresh(db_loc)
    return db_loc

@router.put("/locations/{loc_id}", response_model=schemas.StockLocationResponse)
def update_location(loc_id: int, payload: schemas.StockLocationUpdate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "stock.locations.manage")

    loc = db.query(models.StockLocation).filter(models.StockLocation.id == loc_id).first()
    if not loc:
        raise HTTPException(status_code=404, detail="Emplacement introuvable.")

    data = payload.model_dump(exclude_unset=True)
    if "name" in data:
        name = (data["name"] or "").strip()
        if not name:
            raise HTTPException(status_code=400, detail="Nom d'emplacement obligatoire.")
        existing = (
            db.query(models.StockLocation)
            .filter(
                models.StockLocation.name == name,
                models.StockLocation.parent_id == loc.parent_id,
                models.StockLocation.id != loc_id,
            )
            .first()
        )
        if existing:
            raise HTTPException(status_code=400, detail="Un emplacement porte déjà ce nom.")
        loc.name = name
    if "usage" in data and data["usage"]:
        loc.usage = data["usage"]
    if "parent_id" in data:
        if data["parent_id"] == loc_id:
            raise HTTPException(status_code=400, detail="Un emplacement ne peut pas être son propre parent.")
        if data["parent_id"]:
            parent = db.query(models.StockLocation).filter_by(id=data["parent_id"], is_active=True).first()
            if not parent:
                raise HTTPException(status_code=400, detail="Zone parente introuvable ou archivée.")
            loc.usage = parent.usage
        loc.parent_id = data["parent_id"]
    if "is_active" in data and data["is_active"] is not None:
        loc.is_active = data["is_active"]

    db.commit()
    db.refresh(loc)
    return loc

@router.delete("/locations/{loc_id}")
def delete_location(loc_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "stock.locations.manage")

    loc = db.query(models.StockLocation).filter(models.StockLocation.id == loc_id).first()
    if not loc:
        return {"status": "success"}

    # Protection 1 : Les 4 lieux systèmes vitaux
    if loc.usage in ['supplier', 'customer', 'inventory', 'production'] and loc.parent_id is None:
        raise HTTPException(400, f"Action Interdite : Ce lieu virtuel système ({loc.usage}) est vital pour le moteur à partie double.")

    # Protection 2 : Détection de stock vivant dans l'arbre entier
    def check_active_stock(location_id):
        active_quants = db.query(models.StockQuant).filter(
            models.StockQuant.location_id == location_id, 
            models.StockQuant.quantity > 0
        ).count()
        if active_quants > 0: return True
        
        children = db.query(models.StockLocation).filter(models.StockLocation.parent_id == location_id).all()
        for child in children:
            if check_active_stock(child.id): return True
        return False

    if check_active_stock(loc_id):
        raise HTTPException(400, "Action Interdite : Cet emplacement (ou une de ses étagères) contient du stock actif. Transférez le stock avant la suppression/archivage.")

    # Get all descendants
    def get_all_descendants(location_id):
        children = db.query(models.StockLocation).filter(models.StockLocation.parent_id == location_id).all()
        descendants = [c.id for c in children]
        for c in children:
            descendants.extend(get_all_descendants(c.id))
        return descendants
        
    all_loc_ids = [loc_id] + get_all_descendants(loc_id)

    # Protection 3 : Archivage au lieu de suppression s'il y a un historique dans l'arbre
    historical_moves = db.query(models.StockMove).filter(
        or_(
            models.StockMove.location_id.in_(all_loc_ids),
            models.StockMove.location_dest_id.in_(all_loc_ids)
        )
    ).count()

    if historical_moves > 0:
        # On archive tout l'arbre
        db.query(models.StockLocation).filter(models.StockLocation.id.in_(all_loc_ids)).update({"is_active": False}, synchronize_session=False)
        db.commit()
        return {"status": "archived"}
    else:
        # On supprime physiquement les Quants vides pour éviter une IntegrityError
        db.query(models.StockQuant).filter(models.StockQuant.location_id.in_(all_loc_ids)).delete(synchronize_session=False)
        db.delete(loc)
        db.commit()
        return {"status": "deleted"}

@router.get("/quants", response_model=List[schemas.StockQuantResponse])
def get_all_quants(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    # Renvoie tous les quants dont la quantité est > 0 pour l'affichage Odoo.
    # Sur les emplacements internes, le réservé ferme et le disponible sont
    # calculés PAR EMPLACEMENT (une réservation ancrée ne pèse que sur son
    # emplacement) — les totaux globaux de la variante ne s'appliquent pas à
    # une fiche emplacement et y créeraient du double comptage.
    quants = (
        db.query(models.StockQuant)
        .options(joinedload(models.StockQuant.location))
        .filter(models.StockQuant.quantity > 0)
        .all()
    )
    blind_locations = _active_blind_location_ids(db, user)
    if blind_locations:
        quants = [
            quant for quant in quants
            if quant.location_id not in blind_locations
        ]
    for quant in quants:
        if quant.location and quant.location.usage == "internal":
            _physical, reserved, available = available_quantity_at_location(db, quant.variant_id, quant.location_id)
            quant.reserved_quantity = reserved
            quant.available_quantity = available
    return quants

# --- CHATTER (AUDIT LOG) CHANNELS ---

@router.get("/chatter/{model_name}/{record_id}", response_model=List[schemas.ChatterMessageResponse])
def get_chatter(model_name: str, record_id: int, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    return db.query(models.ChatterMessage).filter(
        models.ChatterMessage.model_name == model_name,
        models.ChatterMessage.record_id == record_id
    ).order_by(models.ChatterMessage.created_at.desc()).all()

@router.post("/chatter", response_model=schemas.ChatterMessageResponse)
def post_chatter_message(msg: schemas.ChatterMessageCreate, db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    db_msg = models.ChatterMessage(
        model_name=msg.model_name,
        record_id=msg.record_id,
        body=msg.body,
        author=user.get("sub", "User"),
        is_system_log=msg.is_system_log
    )
    db.add(db_msg)
    db.commit()
    db.refresh(db_msg)
    return db_msg

# --- BULK IMPORT / EXPORT EXCEL ---

@router.get("/import/template")
def get_import_template():
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Import_PIM"
    headers = [
        "Reference_Parent", "Nom_Famille", "Matiere", "Unite", "Fournisseur", "Type_Article",
        "Visible_Vente_PDV", "Image_Lien", "Fiche_Tech_Lien", "Gammes_Compatibles", "Reference_Variante", "Code_Barre", "Specificites_Couleur",
        "Prix_Achat", "Seuil_Alerte", "Chemin_Rangement"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col).fill = PatternFill(start_color="333333", fill_type="solid")
    
    # Une ligne d'exemple
    ws.append([
        "3186", "Double poignée de porte Cortizo", "QUINCAILLERIE", "pce", "CORTIZO", "stockable",
        "0", "", "", "COR 60, COR 70", "318601", "340001000211", "Blanc",
        25.50, 10, "A1-R3-B"
    ])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=MMG_Template_Import_Produits.xlsx"
    return response

@router.post("/import/upload")
async def upload_import_file(file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")
        
    if not file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Le fichier doit être un .xlsx")

    content = await file.read()
    stream = io.BytesIO(content)
    
    try:
        wb = openpyxl.load_workbook(stream, data_only=True)
        ws = wb.active
    except Exception as e:
        raise HTTPException(400, "Impossible de lire le fichier Excel.")
        
    rows = list(ws.iter_rows(values_only=True))
    if len(rows) < 1:
        raise HTTPException(400, "Le fichier est vide.")
        
    headers = [str(h).strip() if h else "" for h in rows[0]]
    required_cols = {"Reference_Parent", "Nom_Famille", "Reference_Variante"}
    
    if not required_cols.issubset(set(headers)):
        raise HTTPException(400, f"Colonnes manquantes. Requis : {', '.join(required_cols)}")
        
    created_products = 0
    created_variants = 0
    
    # Convert header to row dict mapping
    for i in range(1, len(rows)):
        row_vals = rows[i]
        # Ignore empty rows
        if not any(row_vals):
            continue
            
        row = dict(zip(headers, row_vals))
        
        ref_parent = str(row.get("Reference_Parent") or "").strip()
        nom = str(row.get("Nom_Famille") or "").strip()
        ref_var = str(row.get("Reference_Variante") or "").strip()
        
        if not ref_parent or not nom or not ref_var:
            continue
            
        # 1. Chercher ou créer le Product (Parent)
        product = db.query(models.Product).filter(models.Product.reference_base == ref_parent).first()
        image_url = str(row.get("Image_Lien") or "").strip()
        
        if not product:
            product = models.Product(
                reference_base=ref_parent,
                name=nom,
                material_type=str(row.get("Matiere") or "INCONNU"),
                unit=str(row.get("Unite") or "pce"),
                supplier=str(row.get("Fournisseur") or ""),
                product_type=str(row.get("Type_Article") or "stockable"),
                available_in_pos=True if str(row.get("Visible_Vente_PDV") or "") in ["1", "true", "oui", "OUI", "True", 1] else False,
                image_url=image_url if image_url else None,
                technical_doc_url=str(row.get("Fiche_Tech_Lien") or "").strip() or None,
                compatible_series=str(row.get("Gammes_Compatibles") or "").strip()
            )
            db.add(product)
            db.flush() # get product.id
            created_products += 1
        elif image_url and not product.image_url:
            product.image_url = image_url # Update if provided and missing
            
        # 2. Créer la Variante (Enfant)
        variant = db.query(models.ProductVariant).filter(models.ProductVariant.reference == ref_var).first()
        if not variant:
            code_barre = str(row.get("Code_Barre") or "").strip()
            
            try:
                cp = float(row.get("Prix_Achat") or 0)
            except:
                cp = 0
            
            try:    
                mt = float(row.get("Seuil_Alerte") or 10)
            except:
                mt = 10
                
            variant = models.ProductVariant(
                product_id=product.id,
                reference=ref_var,
                barcode=code_barre if code_barre else None,
                color=str(row.get("Specificites_Couleur") or ""),
                cost_price=cp,
                min_threshold=mt,
                location=str(row.get("Chemin_Rangement") or "").strip()
            )
            db.add(variant)
            created_variants += 1
            
    db.commit()
    return {"message": f"Import réussi : {created_products} familles créées, {created_variants} déclinaisons ajoutées."}

@router.get("/export/inventory")
def export_inventory_xlsx(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    # Document comptable (prix d'achat + valorisation) : même permission que
    # l'export d'audit des mouvements (/transactions/export) — stock.adjust
    # est détenue par ADMIN/MANAGER/CHEF_STOCK dans la matrice.
    _require_permission(db, user, "stock.adjust")
    from sqlalchemy.orm import joinedload
    # Document comptable : seul le stock physiquement détenu (emplacements
    # internes actifs) est exporté. Les emplacements virtuels `customer`
    # (cumul des ventes), `inventory` (pertes/écarts) et `supplier` sont
    # exclus — sinon la valorisation additionne stock réel + vendu + pertes.
    quants = (
        db.query(models.StockQuant)
        .join(models.StockLocation, models.StockQuant.location_id == models.StockLocation.id)
        .options(
            joinedload(models.StockQuant.variant).joinedload(models.ProductVariant.product),
            joinedload(models.StockQuant.location)
        )
        .filter(
            models.StockLocation.usage == "internal",
            models.StockLocation.is_active == True,
        )
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Inventaire_Interne_MMG"
    headers = [
        "Lieu / Magasin (stock interne)", "Reference", "Designation", "Code Barre", "Type",
        "Quantite", "Unite", "Prix Unitaire", "Valeur Totale"
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col).fill = PatternFill(start_color="1E3A8A", fill_type="solid")
        
    for q in quants:
        v = q.variant
        p = v.product if v else None
        loc = q.location
        if not v or not p or not loc:
            continue
            
        valeur_totale = float(q.quantity) * float(v.cost_price or 0)
        ws.append([
            loc.name,
            v.reference,
            f"{p.name} ({v.color or 'Std'})",
            v.barcode or "",
            p.product_type,
            q.quantity,
            p.unit,
            float(v.cost_price or 0),
            valeur_totale
        ])
    
    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=MMG_Inventaire_Actuel.xlsx"
    return response

@router.get("/catalog/drafts/export")
def export_draft_catalog_xlsx(db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    # Pendant export du POST /catalog/drafts/import : même permission
    # catalog.qualify (ADMIN/MANAGER/CHEF_STOCK dans la matrice).
    _require_permission(db, user, "catalog.qualify")
    drafts = (
        db.query(models.Product)
        .options(joinedload(models.Product.variants))
        .filter(models.Product.catalog_status == "DRAFT")
        .order_by(models.Product.supplier.asc(), models.Product.reference_base.asc())
        .all()
    )

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Brouillons_Catalogue"
    headers = [
        "Reference_Parent",
        "Reference_Variante",
        "Fournisseur",
        "Nom_Famille",
        "Categorie",
        "Matiere",
        "Unite",
        "Ref_Fournisseur",
        "Couleur",
        "Finition",
        "Longueur_Unite",
        "Conditionnement",
        "Unites_Conditionnement",
        "Emplacement",
        "Gammes_Compatibles",
        "Statut_Catalogue",
        "Commentaire",
    ]
    ws.append(headers)
    for col in range(1, len(headers) + 1):
        ws.cell(row=1, column=col).font = Font(bold=True, color="FFFFFF")
        ws.cell(row=1, column=col).fill = PatternFill(start_color="D97706", fill_type="solid")

    for product in drafts:
        variants = product.variants or [None]
        for variant in variants:
            ws.append(
                [
                    product.reference_base,
                    variant.reference if variant else "",
                    product.supplier or "",
                    product.name,
                    product.category or "",
                    product.material_type,
                    product.unit,
                    variant.supplier_reference if variant else "",
                    variant.color if variant else "",
                    variant.finish if variant else "",
                    variant.length_per_unit if variant else "",
                    variant.conditioning if variant else "",
                    variant.units_per_package if variant else "",
                    variant.location if variant else "",
                    product.compatible_series or "",
                    product.catalog_status or "DRAFT",
                    "",
                ]
            )

    for column_cells in ws.columns:
        max_length = max(len(str(cell.value or "")) for cell in column_cells)
        ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 14), 45)

    stream = io.BytesIO()
    wb.save(stream)
    stream.seek(0)
    response = StreamingResponse(iter([stream.getvalue()]), media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    response.headers["Content-Disposition"] = "attachment; filename=MMG_Brouillons_Catalogue.xlsx"
    return response

@router.post("/catalog/drafts/import")
async def import_draft_catalog_updates(file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    _require_permission(db, user, "catalog.qualify")
    if not file.filename or not file.filename.endswith(".xlsx"):
        raise HTTPException(status_code=400, detail="Le fichier doit être un .xlsx")

    try:
        wb = openpyxl.load_workbook(io.BytesIO(await file.read()), data_only=True)
        ws = wb.active
    except Exception:
        raise HTTPException(status_code=400, detail="Impossible de lire le fichier Excel.")

    rows = list(ws.iter_rows(values_only=True))
    if not rows:
        raise HTTPException(status_code=400, detail="Le fichier est vide.")

    headers = [str(h).strip() if h else "" for h in rows[0]]
    required = {"Reference_Parent", "Reference_Variante"}
    if not required.issubset(set(headers)):
        raise HTTPException(status_code=400, detail=f"Colonnes manquantes. Requis : {', '.join(sorted(required))}")

    updated_products = 0
    updated_variants = 0
    skipped = 0
    active_values = {"ACTIVE", "ACTIF", "1", "TRUE", "OUI", "YES"}

    for values in rows[1:]:
        if not any(values):
            continue
        row = dict(zip(headers, values))
        ref_parent = str(row.get("Reference_Parent") or "").strip()
        ref_variant = str(row.get("Reference_Variante") or "").strip()
        if not ref_parent and not ref_variant:
            skipped += 1
            continue

        product = db.query(models.Product).filter(models.Product.reference_base == ref_parent).first() if ref_parent else None
        variant = db.query(models.ProductVariant).filter(models.ProductVariant.reference == ref_variant).first() if ref_variant else None
        if not product and variant:
            product = variant.product
        if not product:
            skipped += 1
            continue
        previous_status = product.catalog_status or "DRAFT"

        def text_value(key: str):
            value = row.get(key)
            return str(value).strip() if value is not None else None

        for key, attr in [
            ("Nom_Famille", "name"),
            ("Categorie", "category"),
            ("Matiere", "material_type"),
            ("Unite", "unit"),
            ("Fournisseur", "supplier"),
            ("Gammes_Compatibles", "compatible_series"),
        ]:
            value = text_value(key)
            if value:
                setattr(product, attr, value)

        status = (text_value("Statut_Catalogue") or "").upper()
        updated_products += 1

        if variant:
            supplier_ref = text_value("Ref_Fournisseur")
            location = text_value("Emplacement")
            color = text_value("Couleur")
            finish = text_value("Finition")
            conditioning = text_value("Conditionnement")
            if supplier_ref:
                variant.supplier_reference = supplier_ref
            if location:
                variant.location = location
            if color:
                variant.color = color
            if finish:
                variant.finish = finish
            if conditioning:
                variant.conditioning = conditioning
            length = row.get("Longueur_Unite")
            if length not in (None, ""):
                try:
                    variant.length_per_unit = float(length)
                except (TypeError, ValueError):
                    pass
            units_per_package = row.get("Unites_Conditionnement")
            if units_per_package not in (None, ""):
                try:
                    variant.units_per_package = float(units_per_package)
                except (TypeError, ValueError):
                    pass
            updated_variants += 1

        if status:
            requested_status = "ACTIVE" if status in active_values else (
                status if status in CATALOG_STATUSES else "DRAFT"
            )
            if requested_status == "ACTIVE":
                issues = _activation_issues(product)
                product.catalog_status = "TO_QUALIFY" if issues else "ACTIVE"
            elif requested_status in {"DRAFT", "TO_QUALIFY"}:
                product.catalog_status = requested_status
        _record_product_audit(
            db,
            product_id=product.id,
            variant_id=variant.id if variant else None,
            user=user,
            action="BULK_CATALOG_UPDATE",
            changes={
                "catalog_status": {
                    "before": previous_status,
                    "after": product.catalog_status or previous_status,
                }
            },
            reason=f"Mise à jour depuis {file.filename or 'fichier Excel'}",
        )

    db.commit()
    return {
        "message": f"Mise à jour brouillons : {updated_products} famille(s), {updated_variants} variante(s), {skipped} ligne(s) ignorée(s).",
        "updated_products": updated_products,
        "updated_variants": updated_variants,
        "skipped": skipped,
    }

@router.post("/import-bom/{sale_order_id}")
async def import_bom_for_sale_order(sale_order_id: int, background_tasks: BackgroundTasks, file: UploadFile = File(...), db: Session = Depends(get_db), user: dict = Depends(get_current_user)):
    from ..core.events import EventBus
    _require_permission(db, user, "workshop.reserve_stock")
        
    sale = db.query(models.SaleOrder).filter(models.SaleOrder.id == sale_order_id).first()
    if not sale:
        raise HTTPException(404, "Sale Order introuvable")
    try:
        technical_dossier = technical_dossier_for_sale(db, sale.id)
    except ValueError as exc:
        raise HTTPException(409, str(exc)) from exc
    if technical_dossier:
        raise HTTPException(
            409,
            "Ce devis est rattaché à un dossier technique. "
            "Utilisez le débit validé pour créer une réservation puis un bon de préparation.",
        )

    contents = await file.read()
    try:
        content_str = contents.decode("utf-8")
    except UnicodeDecodeError:
        content_str = contents.decode("latin-1")

    bom_items = parse_bom_file(content_str, file.filename)
    if not bom_items:
        raise HTTPException(400, "Le fichier est vide ou le format n'est pas reconnu.")

    prod_loc = InventoryService.get_or_create_location(db, "Production Ateliers", "production")

    wh_loc = db.query(models.StockLocation).filter(models.StockLocation.usage == "internal").first()
    if not wh_loc:
        raise HTTPException(400, "Aucun entrepôt physique trouvé.")

    processed_count = 0
    not_found_refs = []
    stock_warnings = []

    for item in bom_items:
        ref = item["reference"]
        qty = item["quantity"]
        
        variant = db.query(models.ProductVariant).filter(or_(
            models.ProductVariant.reference == ref,
            models.ProductVariant.barcode == ref,
            models.ProductVariant.supplier_reference == ref
        )).first()

        if not variant:
            not_found_refs.append(ref)
            continue
            
        src_quant = db.query(models.StockQuant).filter_by(variant_id=variant.id, location_id=wh_loc.id).first()
        current_source_qty = float(src_quant.quantity if src_quant else 0)

        # Check for stock warnings (Non-blocking)
        if current_source_qty < qty:
            shortage = qty - current_source_qty
            stock_warnings.append(f"{variant.reference} : manque {shortage} {variant.product.unit if variant.product else 'unités'} en stock.")

        try:
            result = InventoryService.move_stock(
                db,
                variant_id=variant.id,
                source_location_id=wh_loc.id,
                dest_location_id=prod_loc.id,
                quantity=qty,
                reference=f"PROD-{sale.reference}-BOM",
                notes=f"Débit BOM auto ({file.filename})",
                author="Système / Admin",
                source_screen="stock.legacy_bom_import",
                document_type="sale_order",
                document_reference=sale.reference,
                business_reason="Import BOM historique vers production",
                allow_negative_source=True,
            )
        except ValueError as exc:
            status_code = 423 if "Zone gelée" in str(exc) else 400
            raise HTTPException(status_code=status_code, detail=str(exc)) from exc

        if (
            result.previous_source_quantity is not None
            and result.new_source_quantity is not None
            and result.previous_source_quantity > variant.min_threshold
            and result.new_source_quantity <= variant.min_threshold
        ):
            EventBus.on_stock_alert(variant.reference, result.new_source_quantity, background_tasks)
        
        # Log to Chatter
        audit_log = models.ChatterMessage(
            model_name="variant",
            record_id=variant.id,
            body=f"Consommation pour Commande {sale.reference} (Nomenclature {file.filename}). Quantité réservée : {qty}",
            author=user.get("sub", "BE / Admin"),
            is_system_log=True
        )
        db.add(audit_log)
        
        processed_count += 1
        
    sale.status = "READY_FOR_PROD"
    if stock_warnings:
        sale.notes = (sale.notes or "") + f"\n[ALERTE STOCK] {len(stock_warnings)} ruptures potentielles au lancement."
        
    db.commit()

    return {
        "status": "success", 
        "processed_count": processed_count, 
        "not_found": not_found_refs,
        "warnings": stock_warnings
    }
