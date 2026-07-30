#!/usr/bin/env python3
"""Import a real multi-supplier stock workbook into MMG.

Default mode is a dry-run preview. Use --apply to write products, variants,
stock quants and initial inventory moves.
"""

from __future__ import annotations

import argparse
import csv
import json
import sys
from collections import Counter, defaultdict
from dataclasses import asdict, dataclass
from datetime import datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

sys.path.append(str(Path(__file__).resolve().parents[1]))

from backend import models
from backend.database import SessionLocal
from backend.services.stock_service import InventoryService


EXPECTED_HEADERS = ["Réf", "Nom de l'accessoire", "Quant", "Gamme", "iIlustration"]
INVALID_REFERENCE_VALUES = {"/", "-", "x", "xx", "?", "nc", "n/a", "na"}
ISSUE_DECISIONS = {
    "duplicate_reference_conflict": "Arbitrer la bonne désignation puis réimporter.",
    "duplicate_quantity_disagreement": "Contrôler la quantité réelle; le script ne cumule pas.",
    "duplicate_supplier_reference": "Contrôler que le doublon est volontaire.",
    "invalid_reference": "Corriger la référence ou ignorer la ligne.",
    "missing_reference": "Renseigner la référence fournisseur.",
    "missing_designation": "Compléter la désignation produit.",
    "missing_or_invalid_quantity": "Renseigner une quantité réelle; la ligne n'est pas importée.",
}


@dataclass(frozen=True)
class StockRecord:
    row: int
    supplier: str
    reference: str
    designation: str
    quantity: float | None
    gamme: str
    unit: str = "pce"
    category: str = "ACCESSOIRE"
    material_type: str = "ACCESSOIRE"
    product_type: str = "stockable"
    designation_is_placeholder: bool = False
    quantity_is_valid: bool = True


@dataclass(frozen=True)
class StockIssue:
    severity: str
    code: str
    row: int | None
    supplier: str | None
    reference: str | None
    message: str


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split()).strip()


def normalize_supplier(value: Any) -> str:
    supplier = clean_text(value)
    if "/" in supplier:
        supplier = supplier.split("/", 1)[0]
    return supplier.strip().upper()


def normalize_header(value: Any) -> str:
    return clean_text(value).lower().replace("é", "e").replace("è", "e").replace("'", "")


def is_invalid_reference(reference: str) -> bool:
    normalized = reference.strip().lower()
    return normalized in INVALID_REFERENCE_VALUES


def parse_quantity(value: Any) -> float | None:
    if value is None or value == "":
        return None
    if isinstance(value, str):
        value = value.replace(",", ".").strip()
    try:
        return float(value)
    except (TypeError, ValueError):
        return None


def iter_supplier_blocks(rows: list[tuple[Any, ...]], max_column: int) -> list[tuple[int, str]]:
    blocks: list[tuple[int, str]] = []
    current_supplier = ""
    expected_headers = [normalize_header(header) for header in EXPECTED_HEADERS]
    for start in range(0, max_column, 6):
        supplier = normalize_supplier(rows[0][start] if start < len(rows[0]) else None)
        if supplier:
            current_supplier = supplier

        headers = [normalize_header(value) for value in rows[2][start : start + 5]]
        if current_supplier and headers[:5] == expected_headers:
            blocks.append((start, current_supplier))
    return blocks


def duplicate_issue_analysis(records: list[StockRecord]) -> list[StockIssue]:
    issues: list[StockIssue] = []
    groups: dict[tuple[str, str], list[StockRecord]] = defaultdict(list)
    for record in records:
        groups[(record.supplier, record.reference)].append(record)

    for (supplier, reference), group in sorted(groups.items()):
        if len(group) <= 1:
            continue

        names = {record.designation.strip().lower() for record in group if record.designation}
        quantities = {float(record.quantity or 0) for record in group}
        rows = ", ".join(str(record.row) for record in group[:8])
        suffix = "..." if len(group) > 8 else ""

        if len(names) > 1:
            issues.append(
                StockIssue(
                    "error",
                    "duplicate_reference_conflict",
                    None,
                    supplier,
                    reference,
                    f"Référence présente {len(group)} fois avec désignations différentes (lignes {rows}{suffix}); "
                    "groupe ignoré jusqu'à arbitrage manuel.",
                )
            )
        elif len(quantities) > 1:
            issues.append(
                StockIssue(
                    "warning",
                    "duplicate_quantity_disagreement",
                    None,
                    supplier,
                    reference,
                    f"Référence présente {len(group)} fois avec quantités différentes (lignes {rows}{suffix}); "
                    "quantité conservée une seule fois, sans cumul automatique.",
                )
            )
        else:
            issues.append(
                StockIssue(
                    "warning",
                    "duplicate_supplier_reference",
                    None,
                    supplier,
                    reference,
                    f"Référence présente {len(group)} fois; gammes fusionnées, quantité non cumulée.",
                )
            )

    return issues


def parse_workbook(path: Path) -> tuple[list[StockRecord], list[StockIssue]]:
    workbook = load_workbook(path, data_only=True)
    sheet = workbook.active
    rows = list(sheet.iter_rows(values_only=True))
    issues: list[StockIssue] = []

    if len(rows) < 5:
        return [], [
            StockIssue("error", "empty_workbook", None, None, None, "Le classeur ne contient pas assez de lignes.")
        ]

    blocks = iter_supplier_blocks(rows, sheet.max_column)
    if not blocks:
        return [], [
            StockIssue(
                "error",
                "missing_blocks",
                None,
                None,
                None,
                "Aucun bloc fournisseur Réf/Nom/Quant/Gamme/Illustration reconnu.",
            )
        ]

    records: list[StockRecord] = []
    for start, supplier in blocks:
        for row_number, row in enumerate(rows[4:], start=5):
            values = list(row[start : start + 5])
            values += [None] * (5 - len(values))
            ref, designation, quantity_raw, gamme, _illustration = values
            if all(value is None for value in values):
                continue

            reference = clean_text(ref)
            name = clean_text(designation)
            quantity = parse_quantity(quantity_raw)
            gamme_text = clean_text(gamme)

            if not reference:
                issues.append(
                    StockIssue(
                        "error",
                        "missing_reference",
                        row_number,
                        supplier,
                        None,
                        "Ligne ignorée: référence manquante.",
                    )
                )
                continue
            if is_invalid_reference(reference):
                issues.append(
                    StockIssue(
                        "error",
                        "invalid_reference",
                        row_number,
                        supplier,
                        reference,
                        "Ligne ignorée: référence non exploitable.",
                    )
                )
                continue
            designation_is_placeholder = not name
            if designation_is_placeholder:
                issues.append(
                    StockIssue(
                        "warning",
                        "missing_designation",
                        row_number,
                        supplier,
                        reference,
                        "Désignation vide; le nom produit utilisera la référence.",
                    )
                )
                name = reference
            quantity_is_valid = quantity is not None
            if not quantity_is_valid:
                issues.append(
                    StockIssue(
                        "warning",
                        "missing_or_invalid_quantity",
                        row_number,
                        supplier,
                        reference,
                        f"Quantité vide ou invalide ({quantity_raw!r}); ligne exclue de l'import.",
                    )
                )
                quantity = 0.0

            records.append(
                StockRecord(
                    row=row_number,
                    supplier=supplier,
                    reference=reference,
                    designation=name,
                    quantity=quantity,
                    gamme=gamme_text,
                    designation_is_placeholder=designation_is_placeholder,
                    quantity_is_valid=quantity_is_valid,
                )
            )

    issues.extend(duplicate_issue_analysis(records))

    return records, issues


def consolidate_records(records: list[StockRecord]) -> list[StockRecord]:
    conflict_keys = {
        (issue.supplier, issue.reference)
        for issue in duplicate_issue_analysis(records)
        if issue.code == "duplicate_reference_conflict"
    }
    consolidated: dict[tuple[str, str], StockRecord] = {}
    for record in records:
        if not record.quantity_is_valid:
            continue
        key = (record.supplier, record.reference)
        if key in conflict_keys:
            continue
        existing = consolidated.get(key)
        if not existing:
            consolidated[key] = record
            continue

        gammes = [value for value in (existing.gamme, record.gamme) if value]
        merged_gamme = ", ".join(dict.fromkeys(part.strip() for gamme in gammes for part in gamme.split(",") if part.strip()))
        existing_qty = float(existing.quantity or 0)
        record_qty = float(record.quantity or 0)
        quantity = existing_qty if existing_qty > 0 else record_qty

        consolidated[key] = StockRecord(
            row=existing.row,
            supplier=existing.supplier,
            reference=existing.reference,
            designation=existing.designation or record.designation,
            quantity=quantity,
            gamme=merged_gamme,
            unit=existing.unit,
            category=existing.category,
            material_type=existing.material_type,
            product_type=existing.product_type,
            designation_is_placeholder=(
                existing.designation_is_placeholder and record.designation_is_placeholder
            ),
            quantity_is_valid=True,
        )
    return list(consolidated.values())


def records_by_supplier_reference(records: list[StockRecord]) -> dict[tuple[str, str], list[StockRecord]]:
    grouped: dict[tuple[str, str], list[StockRecord]] = defaultdict(list)
    for record in records:
        grouped[(record.supplier, record.reference)].append(record)
    return grouped


def build_summary(records: list[StockRecord], issues: list[StockIssue]) -> dict[str, Any]:
    consolidated = consolidate_records(records)
    supplier_counts = Counter(record.supplier for record in consolidated)
    issue_counts = Counter(issue.code for issue in issues)
    positive_count = sum(1 for record in consolidated if (record.quantity or 0) > 0)
    zero_count = sum(1 for record in consolidated if (record.quantity or 0) == 0)
    return {
        "raw_records": len(records),
        "importable_records": len(consolidated),
        "positive_quantity_records": positive_count,
        "zero_quantity_records": zero_count,
        "suppliers": dict(sorted(supplier_counts.items())),
        "issues": dict(sorted(issue_counts.items())),
    }


def build_issue_report_rows(records: list[StockRecord], issues: list[StockIssue]) -> list[dict[str, Any]]:
    grouped = records_by_supplier_reference(records)
    rows: list[dict[str, Any]] = []

    for issue in issues:
        related_records = grouped.get((issue.supplier or "", issue.reference or ""), [])
        if not related_records:
            rows.append(
                {
                    "severity": issue.severity,
                    "code": issue.code,
                    "supplier": issue.supplier or "",
                    "reference": issue.reference or "",
                    "source_row": issue.row or "",
                    "designation": "",
                    "quantity": "",
                    "gamme": "",
                    "decision": ISSUE_DECISIONS.get(issue.code, "Contrôler la ligne."),
                    "message": issue.message,
                }
            )
            continue

        for record in related_records:
            rows.append(
                {
                    "severity": issue.severity,
                    "code": issue.code,
                    "supplier": record.supplier,
                    "reference": record.reference,
                    "source_row": record.row,
                    "designation": record.designation,
                    "quantity": record.quantity if record.quantity is not None else "",
                    "gamme": record.gamme,
                    "decision": ISSUE_DECISIONS.get(issue.code, "Contrôler la ligne."),
                    "message": issue.message,
                }
            )

    return rows


def write_issues_csv(path: Path, records: list[StockRecord], issues: list[StockIssue]) -> None:
    fieldnames = [
        "severity",
        "code",
        "supplier",
        "reference",
        "source_row",
        "designation",
        "quantity",
        "gamme",
        "decision",
        "message",
    ]
    rows = build_issue_report_rows(records, issues)
    with path.open("w", newline="", encoding="utf-8-sig") as file:
        writer = csv.DictWriter(file, fieldnames=fieldnames, delimiter=";")
        writer.writeheader()
        writer.writerows(rows)


def compare_database(records: list[StockRecord]) -> dict[str, Any]:
    consolidated = consolidate_records(records)
    db = SessionLocal()
    try:
        planned_refs = {f"{record.supplier}:{record.reference}" for record in consolidated}
        existing_products = {
            product.reference_base
            for product in db.query(models.Product.reference_base).all()
            if product.reference_base
        }
        existing_variants = {
            variant.reference
            for variant in db.query(models.ProductVariant.reference).all()
            if variant.reference
        }
        existing_refs = existing_products | existing_variants
        matching_refs = sorted(planned_refs & existing_refs)
        return {
            "planned_records": len(planned_refs),
            "existing_matches": len(matching_refs),
            "new_references": len(planned_refs - existing_refs),
            "existing_not_in_file": len(existing_refs - planned_refs),
            "matching_reference_sample": matching_refs[:25],
        }
    finally:
        db.close()


def get_or_create_location(db, name: str) -> models.StockLocation:
    location = db.query(models.StockLocation).filter_by(name=name, usage="internal").first()
    if location:
        return location

    location = models.StockLocation(name=name, usage="internal", is_active=True)
    db.add(location)
    db.flush()
    return location


def import_records(records: list[StockRecord], location_name: str, dry_run: bool, source_document: str = "import_real_stock") -> dict[str, int]:
    consolidated = consolidate_records(records)
    stats = {
        "created_products": 0,
        "updated_products": 0,
        "created_variants": 0,
        "updated_variants": 0,
        "created_quants": 0,
        "updated_quants": 0,
        "created_moves": 0,
        "skipped_conflicting_records": len(records) - len(consolidated),
        "skipped_invalid_quantity_records": sum(
            1 for record in records if not record.quantity_is_valid
        ),
    }

    if dry_run:
        return stats

    db = SessionLocal()
    try:
        location = get_or_create_location(db, location_name)
        inventory_location = InventoryService.get_or_create_location(db, "Virtual/Inventory", "inventory")
        now_ref = datetime.utcnow().strftime("%Y%m%d%H%M%S")

        for record in consolidated:
            product_reference = f"{record.supplier}:{record.reference}"
            variant_reference = product_reference

            product = db.query(models.Product).filter_by(reference_base=product_reference).first()
            if not product:
                product = models.Product(
                    reference_base=product_reference,
                    name=record.designation,
                    category=record.category,
                    material_type=record.material_type,
                    unit=record.unit,
                    supplier=record.supplier,
                    product_type=record.product_type,
                    compatible_series=record.gamme,
                )
                db.add(product)
                db.flush()
                stats["created_products"] += 1
            else:
                if not record.designation_is_placeholder:
                    product.name = record.designation or product.name
                product.category = record.category
                product.material_type = record.material_type
                product.unit = record.unit
                product.supplier = record.supplier
                product.product_type = record.product_type
                product.compatible_series = record.gamme or product.compatible_series
                stats["updated_products"] += 1

            variant = db.query(models.ProductVariant).filter_by(reference=variant_reference).first()
            if not variant:
                variant = models.ProductVariant(
                    product_id=product.id,
                    reference=variant_reference,
                    supplier_reference=record.reference,
                    quantity_in_stock=0,
                    min_threshold=0,
                    location=location_name,
                )
                db.add(variant)
                db.flush()
                stats["created_variants"] += 1
            else:
                variant.product_id = product.id
                variant.supplier_reference = record.reference
                variant.location = location_name
                stats["updated_variants"] += 1

            quant = db.query(models.StockQuant).filter_by(variant_id=variant.id, location_id=location.id).first()
            previous_qty = float(quant.quantity if quant else 0)
            target_qty = float(record.quantity or 0)
            if quant:
                stats["updated_quants"] += 1
            else:
                InventoryService.get_or_create_quant(db, variant.id, location.id)
                stats["created_quants"] += 1

            delta = target_qty - previous_qty
            if abs(delta) > 1e-9:
                source_location_id = inventory_location.id if delta > 0 else location.id
                dest_location_id = location.id if delta > 0 else inventory_location.id
                InventoryService.move_stock(
                    db,
                    variant_id=variant.id,
                    quantity=abs(delta),
                    source_location_id=source_location_id,
                    dest_location_id=dest_location_id,
                    reference=f"INIT-STOCK-{now_ref}",
                    notes=f"Import stock réel fournisseur {record.supplier}",
                    author="Import stock réel",
                    source_screen="scripts/import_real_stock.py",
                    document_type="stock_import",
                    document_reference=source_document,
                    business_reason="Initialisation stock réel validée",
                    allow_negative_source=True,
                )
                stats["created_moves"] += 1

        db.commit()
        return stats
    except Exception:
        db.rollback()
        raise
    finally:
        db.close()


def write_json(
    path: Path,
    records: list[StockRecord],
    issues: list[StockIssue],
    summary: dict[str, Any],
    db_comparison: dict[str, Any] | None = None,
) -> None:
    payload = {
        "summary": summary,
        "db_comparison": db_comparison,
        "issues": [asdict(issue) for issue in issues],
        "records": [asdict(record) for record in consolidate_records(records)],
    }
    path.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")


def main() -> int:
    parser = argparse.ArgumentParser(description="Prévisualise ou importe le stock réel Excel multi-fournisseurs MMG.")
    parser.add_argument("file", type=Path, help="Fichier .xlsx de stock réel.")
    parser.add_argument("--apply", action="store_true", help="Importe réellement en base. Sans cette option: prévisualisation.")
    parser.add_argument("--location", default="WH/Stock", help="Emplacement interne cible pour les quantités.")
    parser.add_argument("--json-out", type=Path, help="Écrit la prévisualisation complète en JSON.")
    parser.add_argument("--issues-csv", type=Path, help="Écrit un rapport CSV des anomalies à traiter.")
    parser.add_argument("--compare-db", action="store_true", help="Compare les références du fichier avec la base active.")
    parser.add_argument("--fail-on-errors", action="store_true", help="Retourne 1 si des erreurs bloquantes sont détectées.")
    parser.add_argument("--allow-errors", action="store_true", help="Autorise --apply malgré des erreurs bloquantes ignorées.")
    args = parser.parse_args()

    if not args.file.exists():
        print(f"FAIL: fichier introuvable: {args.file}", file=sys.stderr)
        return 1

    records, issues = parse_workbook(args.file)
    summary = build_summary(records, issues)
    db_comparison = compare_database(records) if args.compare_db else None
    blocking_errors = [issue for issue in issues if issue.severity == "error"]

    print("# Prévisualisation import stock réel")
    print(json.dumps(summary, ensure_ascii=False, indent=2))
    if db_comparison:
        print("# Comparaison base active")
        print(json.dumps(db_comparison, ensure_ascii=False, indent=2))

    for issue in issues[:30]:
        location = f"ligne {issue.row}" if issue.row else "global"
        print(f"{issue.severity.upper()}: {issue.code} ({location}, {issue.supplier or '-'}/{issue.reference or '-'}) - {issue.message}")
    if len(issues) > 30:
        print(f"INFO: {len(issues) - 30} autres alertes non affichées. Utiliser --json-out pour le détail.")

    if args.json_out:
        write_json(args.json_out, records, issues, summary, db_comparison)
        print(f"JSON: {args.json_out}")
    if args.issues_csv:
        write_issues_csv(args.issues_csv, records, issues)
        print(f"ISSUES_CSV: {args.issues_csv}")

    if args.fail_on_errors and blocking_errors:
        return 1

    if args.apply and blocking_errors and not args.allow_errors:
        print("FAIL: erreurs bloquantes détectées. Corriger le fichier ou relancer avec --allow-errors.", file=sys.stderr)
        return 1

    stats = import_records(records, args.location, dry_run=not args.apply, source_document=args.file.name)
    if args.apply:
        print("# Import appliqué")
    else:
        print("# Dry-run: aucune écriture en base")
    print(json.dumps(stats, ensure_ascii=False, indent=2))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
