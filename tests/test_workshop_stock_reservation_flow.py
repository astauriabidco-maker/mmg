import io
import json
from datetime import datetime

from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import sessionmaker
from sqlalchemy.pool import StaticPool

from backend import database, models
from backend.core import security
from backend.main import app


def _auth_headers(session_factory, username: str, role: str = "ADMIN") -> dict:
    """Crée l'utilisateur en base si besoin, puis émet un JWT valide pour lui."""
    with session_factory() as db:
        if not db.query(models.User).filter(models.User.username == username).first():
            db.add(models.User(username=username, pin_hash="test-pin", role=role, is_active=True))
            db.commit()
    token = security.create_access_token({"sub": username, "role": role})
    return {"Authorization": f"Bearer {token}"}


def _prepare_and_handover(client: TestClient, headers: dict, reservation: dict) -> dict:
    create_response = client.post(
        "/v2/stock/workshop-preparations",
        headers=headers,
        json={"reservation_id": reservation["id"]},
    )
    assert create_response.status_code == 200, create_response.text
    preparation = create_response.json()
    for line in preparation["lines"]:
        update_response = client.patch(
            f"/v2/stock/workshop-preparations/{preparation['id']}/lines/{line['id']}",
            headers=headers,
            json={"prepared_quantity": line["planned_quantity"]},
        )
        assert update_response.status_code == 200, update_response.text
        preparation = update_response.json()
    handover_response = client.post(
        f"/v2/stock/workshop-preparations/{preparation['id']}/handover",
        headers=headers,
    )
    assert handover_response.status_code == 200, handover_response.text
    return preparation


def test_workshop_debit_reservation_is_consumed_only_when_confirmed():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            product = models.Product(
                reference_base="SEPALUMIC:7007",
                name="Bavette de faitage",
                material_type="ALU",
                unit="barre",
                supplier="SEPALUMIC",
                product_type="stockable",
            )
            db.add(product)
            db.flush()
            variant = models.ProductVariant(
                product_id=product.id,
                reference="SEPALUMIC:7007",
                supplier_reference="7007",
                quantity_in_stock=5,
                min_threshold=0,
            )
            location = models.StockLocation(name="WH/Stock", usage="internal", is_active=True)
            db.add_all([variant, location])
            db.flush()
            db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=5))
            sale = models.SaleOrder(
                reference="DEV-ATELIER-1",
                client_name="Client atelier",
                status="IN_DESIGN",
                tax_rate=20,
            )
            db.add(sale)
            db.flush()
            db.add(
                models.SaleOrderLine(
                    order_id=sale.id,
                    description="Menuiserie ALU Sepalumic",
                    quantity=1,
                    unit_price=1000,
                )
            )
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        preview_response = client.post(
            "/v2/stock/workshop-debits/preview",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert preview_response.status_code == 200, preview_response.text
        assert preview_response.json()["summary"]["stock_match_status"] == {"ok": 1}
        assert preview_response.json()["issues"] == []

        reserve_response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert reserve_response.status_code == 200, reserve_response.text
        reservation = reserve_response.json()
        assert reservation["status"] == "reserved"
        assert reservation["sale_order_id"] == sale_id
        assert reservation["sale_status"] == "READY_FOR_PROD"
        assert reservation["sale_reference"] == "DEV-ATELIER-1"
        assert reservation["lines"][0]["reserved_quantity"] == 3

        with TestingSessionLocal() as db:
            quant = db.query(models.StockQuant).one()
            variant = db.query(models.ProductVariant).one()
            assert quant.quantity == 5
            assert variant.quantity_in_stock == 5

        _prepare_and_handover(client, headers, reservation)

        with TestingSessionLocal() as db:
            source_quant = (
                db.query(models.StockQuant)
                .join(models.StockLocation)
                .filter(models.StockLocation.name == "WH/Stock")
                .one()
            )
            staging_quant = (
                db.query(models.StockQuant)
                .join(models.StockLocation)
                .filter(models.StockLocation.name == "ATELIER/Préparation")
                .one()
            )
            assert source_quant.quantity == 2
            assert staging_quant.quantity == 3

        early_consume_response = client.post(
            f"/v2/stock/workshop-debits/reservations/{reservation['id']}/consume",
            headers=headers,
        )
        assert early_consume_response.status_code == 400
        assert "Lancez la fabrication" in early_consume_response.text

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).filter(models.SaleOrder.id == sale_id).one()
            sale_db.status = "IN_PRODUCTION"
            db.commit()

        consume_response = client.post(
            f"/v2/stock/workshop-debits/reservations/{reservation['id']}/consume",
            headers=headers,
        )
        assert consume_response.status_code == 200, consume_response.text
        assert consume_response.json()["created_moves"] == 1

        with TestingSessionLocal() as db:
            source_quant = db.query(models.StockQuant).join(models.StockLocation).filter(models.StockLocation.name == "WH/Stock").one()
            dest_quant = (
                db.query(models.StockQuant)
                .join(models.StockLocation)
                .filter(models.StockLocation.name == "Production Ateliers")
                .one()
            )
            staging_quant = (
                db.query(models.StockQuant)
                .join(models.StockLocation)
                .filter(models.StockLocation.name == "ATELIER/Préparation")
                .one()
            )
            variant = db.query(models.ProductVariant).one()
            moves = db.query(models.StockMove).order_by(models.StockMove.id).all()
            reservation_db = db.query(models.StockReservation).one()
            preparation_db = db.query(models.WorkshopPreparation).one()

        assert source_quant.quantity == 2
        assert staging_quant.quantity == 0
        assert dest_quant.quantity == 3
        assert variant.quantity_in_stock == 2
        assert len(moves) == 2
        assert moves[0].reference.startswith("BPA-")
        assert moves[1].reference.startswith("DEBIT-ATELIER-")
        assert "Débit atelier réel" in moves[1].notes
        assert "DEV-ATELIER-1" in moves[1].notes
        assert reservation_db.status == "consumed"
        assert preparation_db.status == "consumed"
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_workshop_debit_reservation_can_be_cancelled_without_stock_move():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            product = models.Product(
                reference_base="SEPALUMIC:7007",
                name="Bavette de faitage",
                material_type="ALU",
                unit="barre",
                supplier="SEPALUMIC",
                product_type="stockable",
            )
            db.add(product)
            db.flush()
            variant = models.ProductVariant(
                product_id=product.id,
                reference="SEPALUMIC:7007",
                supplier_reference="7007",
                quantity_in_stock=5,
                min_threshold=0,
            )
            location = models.StockLocation(name="WH/Stock", usage="internal", is_active=True)
            db.add_all([variant, location])
            db.flush()
            db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=5))
            sale = models.SaleOrder(
                reference="DEV-ATELIER-CANCEL",
                client_name="Client atelier",
                status="VALIDATED",
                tax_rate=20,
            )
            db.add(sale)
            db.flush()
            db.add(
                models.SaleOrderLine(
                    order_id=sale.id,
                    description="Menuiserie ALU Sepalumic",
                    quantity=1,
                    unit_price=1000,
                )
            )
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        reserve_response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert reserve_response.status_code == 200, reserve_response.text
        reservation = reserve_response.json()

        cancel_response = client.post(
            f"/v2/stock/workshop-debits/reservations/{reservation['id']}/cancel",
            headers=headers,
        )
        assert cancel_response.status_code == 200, cancel_response.text
        assert cancel_response.json()["cancelled_lines"] == 1
        assert cancel_response.json()["released_quantity"] == 3

        with TestingSessionLocal() as db:
            quant = db.query(models.StockQuant).one()
            variant = db.query(models.ProductVariant).one()
            reservation_db = db.query(models.StockReservation).one()
            line = db.query(models.StockReservationLine).one()
            moves = db.query(models.StockMove).count()

        assert quant.quantity == 5
        assert variant.quantity_in_stock == 5
        assert reservation_db.status == "cancelled"
        assert line.status == "cancelled"
        assert moves == 0
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_sale_workshop_preparation_reserves_stock_without_physical_debit():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            product = models.Product(
                reference_base="CORTIZO:2000",
                name="Profil Cortizo 2000",
                material_type="ALU",
                unit="barre",
                supplier="CORTIZO",
                product_type="stockable",
            )
            db.add(product)
            db.flush()
            variant = models.ProductVariant(
                product_id=product.id,
                reference="CORTIZO:2000",
                supplier_reference="2000",
                quantity_in_stock=5,
                min_threshold=0,
            )
            location = models.StockLocation(name="WH/Stock", usage="internal", is_active=True)
            db.add_all([variant, location])
            db.flush()
            db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=5))
            sale = models.SaleOrder(
                reference="DEV-PREP-ATELIER",
                client_name="Client préparation",
                status="VALIDATED",
                workflow_type="FABRICATION_FROM_MEASURE",
                tax_rate=20,
            )
            db.add(sale)
            db.flush()
            db.add(
                models.SaleOrderLine(
                    order_id=sale.id,
                    description="Menuiserie ALU Cortizo",
                    quantity=1,
                    unit_price=1000,
                )
            )
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")
        content = b"CORTIZO GAMME BASE\r\nVER TEST\r\nRAL;2000;PROFIL;4;barre  6,50\r\n"

        preview_response = client.post(
            f"/v2/sales/{sale_id}/prepare-workshop/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert preview_response.status_code == 200, preview_response.text
        assert preview_response.json()["summary"]["stock_match_status"] == {"ok": 1}

        reserve_response = client.post(
            f"/v2/sales/{sale_id}/prepare-workshop/reserve",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert reserve_response.status_code == 200, reserve_response.text
        assert reserve_response.json()["status"] == "READY_FOR_PROD"
        assert reserve_response.json()["reserved_lines"] == 1

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
            reservation = db.query(models.StockReservation).one()
            quant = db.query(models.StockQuant).one()
            variant = db.query(models.ProductVariant).one()
            moves = db.query(models.StockMove).count()

        assert sale_db.status == "READY_FOR_PROD"
        assert reservation.sale_order_id == sale_id
        assert reservation.status == "reserved"
        assert quant.quantity == 5
        assert variant.quantity_in_stock == 5
        assert moves == 0
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_launch_production_is_idempotent_and_links_reservation_to_order():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-LAUNCH-ATELIER",
                client_name="Client lancement",
                status="READY_FOR_PROD",
                workflow_type="FABRICATION_FROM_MEASURE",
                tax_rate=20,
            )
            db.add(sale)
            db.flush()
            line = models.SaleOrderLine(
                order_id=sale.id,
                description="Chassis ALU Cortizo",
                quantity=2,
                unit_price=1000,
                visual_config=json.dumps(
                    {
                        "type": "Coulissant 2 vantaux",
                        "width": 1800,
                        "height": 2150,
                        "material": "ALU",
                        "color": "RAL 7016",
                    }
                ),
            )
            db.add(line)
            reservation = models.StockReservation(
                reference="RSV-LAUNCH-1",
                sale_order_id=sale.id,
                status="reserved",
                source_label="SEPVER.TXT",
                created_by="test",
            )
            db.add(reservation)
            db.flush()
            db.add(
                models.StockReservationLine(
                    reservation_id=reservation.id,
                    supplier="CORTIZO",
                    supplier_reference="2000",
                    designation="Profil atelier",
                    unit="barre",
                    requested_quantity=1,
                    reserved_quantity=1,
                    status="reserved",
                )
            )
            db.commit()
            sale_id = sale.id
            line_id = line.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")

        first_response = client.post(f"/v2/sales/{sale_id}/launch-production", headers=headers)
        assert first_response.status_code == 200, first_response.text
        assert first_response.json()["created_orders"] == 1
        assert first_response.json()["linked_reservations"] == 1

        second_response = client.post(f"/v2/sales/{sale_id}/launch-production", headers=headers)
        assert second_response.status_code == 200, second_response.text
        assert second_response.json()["created_orders"] == 0
        assert second_response.json()["existing_orders"] == 1

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
            order = db.query(models.Order).one()
            plans = db.query(models.Planning).all()
            reservation_db = db.query(models.StockReservation).one()

        assert sale_db.status == "IN_PRODUCTION"
        assert order.sale_order_id == sale_id
        assert order.sale_order_line_id == line_id
        assert order.width == 1800
        assert order.height == 2150
        assert order.quantity == 2
        assert order.reference.startswith("PROD-LAUNCH-ATELIER-L")
        assert len(plans) == 1
        assert reservation_db.production_order_id == order.id
        assert reservation_db.order_reference == order.reference
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_sale_workshop_preview_rejects_wrong_sales_status():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-FAB-DRAFT",
                client_name="Client fabrication brouillon",
                status="DRAFT",
                workflow_type="FABRICATION_FROM_MEASURE",
                tax_rate=20,
            )
            db.add(sale)
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")
        content = b"CORTIZO GAMME BASE\r\nVER TEST\r\nRAL;2000;PROFIL;4;barre  6,50\r\n"

        response = client.post(
            f"/v2/sales/{sale_id}/prepare-workshop/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )

        assert response.status_code == 400
        assert "Préparation atelier autorisée uniquement" in response.text
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_launch_production_requires_active_workshop_reservation():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-FAB-SANS-RESERVATION",
                client_name="Client sans réservation",
                status="READY_FOR_PROD",
                workflow_type="FABRICATION_FROM_MEASURE",
                tax_rate=20,
            )
            db.add(sale)
            db.flush()
            db.add(
                models.SaleOrderLine(
                    order_id=sale.id,
                    description="Chassis ALU",
                    quantity=1,
                    unit_price=1000,
                    visual_config=json.dumps(
                        {
                            "type": "Fenêtre",
                            "width": 1200,
                            "height": 1400,
                            "material": "ALU",
                        }
                    ),
                )
            )
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")

        response = client.post(f"/v2/sales/{sale_id}/launch-production", headers=headers)

        assert response.status_code == 400
        assert "réservation atelier active" in response.text

        with TestingSessionLocal() as db:
            assert db.query(models.Order).count() == 0
            assert db.query(models.Planning).count() == 0
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_sale_workshop_preparation_rejects_free_sale_quote():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-LIBRE",
                client_name="Client pièces",
                status="VALIDATED",
                workflow_type="FREE_SALE",
                tax_rate=20,
            )
            db.add(sale)
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")
        content = b"CORTIZO GAMME BASE\r\nVER TEST\r\nRAL;2000;PROFIL;4;barre  6,50\r\n"

        response = client.post(
            f"/v2/sales/{sale_id}/prepare-workshop/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert response.status_code == 400
        assert "devis libre" in response.text
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_sale_workshop_preparation_requires_measure_for_fabrication_estimate():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-PRE-FAB",
                client_name="Client fabrication",
                status="VALIDATED",
                workflow_type="FABRICATION_ESTIMATE",
                tax_rate=20,
            )
            db.add(sale)
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")
        content = b"CORTIZO GAMME BASE\r\nVER TEST\r\nRAL;2000;PROFIL;4;barre  6,50\r\n"

        response = client.post(
            f"/v2/sales/{sale_id}/prepare-workshop/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert response.status_code == 400
        assert "métré" in response.text
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_create_measure_from_free_sale_marks_it_as_fabrication_estimate():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            sale = models.SaleOrder(
                reference="DEV-LIBRE-A-CONVERTIR",
                client_name="Client conversion",
                status="VALIDATED",
                workflow_type="FREE_SALE",
                tax_rate=20,
            )
            db.add(sale)
            db.commit()
            sale_id = sale.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "sales-manager")
        response = client.post(f"/v2/mmg/from-sale/{sale_id}", headers=headers)
        assert response.status_code == 200, response.text

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
            dossier = db.query(models.MMG).one()

        assert sale_db.workflow_type == "FABRICATION_ESTIMATE"
        assert dossier.sale_order_id == sale_id
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_workshop_reservation_requires_validated_and_coherent_sale_order():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            draft = models.SaleOrder(reference="DEV-DRAFT", client_name="Draft", status="DRAFT", tax_rate=20)
            pvc = models.SaleOrder(reference="DEV-PVC", client_name="PVC", status="VALIDATED", tax_rate=20)
            db.add_all([draft, pvc])
            db.flush()
            db.add_all(
                [
                    models.SaleOrderLine(order_id=draft.id, description="Menuiserie ALU", quantity=1, unit_price=1000),
                    models.SaleOrderLine(order_id=pvc.id, description="Fenêtre PVC", quantity=1, unit_price=1000),
                ]
            )
            db.commit()
            draft_id = draft.id
            pvc_id = pvc.id

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        draft_response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(draft_id), "allow_missing": "true"},
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert draft_response.status_code == 400
        assert "réservation autorisée seulement après validation" in draft_response.text

        mismatch_response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(pvc_id), "allow_missing": "true"},
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert mismatch_response.status_code == 400
        assert "Incohérence matière" in mismatch_response.text
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_workshop_debit_contexts_include_active_production_orders_without_sales():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        with TestingSessionLocal() as db:
            order = models.Order(
                reference="CMD-ATELIER-1",
                width=1200,
                height=900,
                material=models.MaterialType.ALU,
                client_name="Client atelier",
                quantity=1,
            )
            db.add(order)
            db.flush()
            db.add(
                models.Planning(
                    order_id=order.id,
                    station="ALU_DEBIT",
                    status=models.PlanningStatus.PENDING,
                )
            )
            db.commit()

        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")

        response = client.get("/v2/stock/workshop-debits/contexts", headers=headers)

        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["sales"] == []
        assert payload["production_orders"] == [
            {
                "type": "production_order",
                "id": 1,
                "reference": "CMD-ATELIER-1",
                "client_name": "Client atelier",
                "status": "PENDING",
                "material": "ALU",
                "station": "ALU_DEBIT",
                "label": "CMD-ATELIER-1 - Client atelier (ALU)",
            }
        ]
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_workshop_debit_preview_is_allowed_without_context_but_flags_workflow():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        response = client.post(
            "/v2/stock/workshop-debits/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )

        assert response.status_code == 200, response.text
        payload = response.json()
        assert payload["summary"]["debit_lines"] == 1
        assert any(issue["code"] == "missing_workflow_context" for issue in payload["issues"])
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_workshop_unknown_lines_can_create_zero_stock_draft_products():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        create_response = client.post(
            "/v2/stock/workshop-debits/draft-products",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )

        assert create_response.status_code == 200, create_response.text
        assert create_response.json()["created"] == 1
        assert create_response.json()["references"] == ["SEPALUMIC:7007"]

        preview_response = client.post(
            "/v2/stock/workshop-debits/preview",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )

        assert preview_response.status_code == 200, preview_response.text
        payload = preview_response.json()
        assert payload["summary"]["stock_match_status"] == {"shortage": 1}

        with TestingSessionLocal() as db:
            product = db.query(models.Product).one()
            variant = db.query(models.ProductVariant).one()
            quant = db.query(models.StockQuant).one()

        assert product.name.startswith("[BROUILLON]")
        assert product.supplier == "SEPALUMIC"
        assert product.material_type == "ALU"
        assert product.catalog_status == "DRAFT"
        assert variant.reference == "SEPALUMIC:7007"
        assert variant.supplier_reference == "7007"
        assert variant.quantity_in_stock == 0
        assert quant.quantity == 0
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def test_draft_catalog_can_be_exported_and_reimported():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db

    try:
        client = TestClient(app)
        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        content = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"

        create_response = client.post(
            "/v2/stock/workshop-debits/draft-products",
            headers=headers,
            files=[("files", ("SEPVER.TXT", content, "text/plain"))],
        )
        assert create_response.status_code == 200, create_response.text

        export_response = client.get("/v2/stock/catalog/drafts/export", headers=headers)
        assert export_response.status_code == 200, export_response.text

        workbook = load_workbook(io.BytesIO(export_response.content))
        sheet = workbook.active
        headers_row = [cell.value for cell in sheet[1]]
        values = {header: index + 1 for index, header in enumerate(headers_row)}
        sheet.cell(row=2, column=values["Nom_Famille"]).value = "Profil SEPALUMIC 7007 - Bavette de faitage"
        sheet.cell(row=2, column=values["Longueur_Unite"]).value = 6500
        sheet.cell(row=2, column=values["Emplacement"]).value = "Rack ALU A1"
        sheet.cell(row=2, column=values["Statut_Catalogue"]).value = "ACTIVE"
        buffer = io.BytesIO()
        workbook.save(buffer)
        buffer.seek(0)

        import_response = client.post(
            "/v2/stock/catalog/drafts/import",
            headers=headers,
            files=[("file", ("drafts.xlsx", buffer.getvalue(), "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"))],
        )
        assert import_response.status_code == 200, import_response.text
        assert import_response.json()["updated_products"] == 1
        assert import_response.json()["updated_variants"] == 1

        with TestingSessionLocal() as db:
            product = db.query(models.Product).one()
            variant = db.query(models.ProductVariant).one()

        assert product.name == "Profil SEPALUMIC 7007 - Bavette de faitage"
        assert product.catalog_status == "ACTIVE"
        assert variant.length_per_unit == 6500
        assert variant.location == "Rack ALU A1"
        assert variant.quantity_in_stock == 0
    finally:
        app.dependency_overrides.pop(database.get_db, None)
        models.Base.metadata.drop_all(bind=engine)


def _seed_stock_and_sale(db, sale_status, reference, workflow_type=None, with_visual_config=False):
    """Crée un produit/variant SEPALUMIC:7007 en stock et un devis au statut donné."""
    product = models.Product(
        reference_base="SEPALUMIC:7007",
        name="Bavette de faitage",
        material_type="ALU",
        unit="barre",
        supplier="SEPALUMIC",
        product_type="stockable",
    )
    db.add(product)
    db.flush()
    variant = models.ProductVariant(
        product_id=product.id,
        reference="SEPALUMIC:7007",
        supplier_reference="7007",
        quantity_in_stock=5,
        min_threshold=0,
    )
    location = models.StockLocation(name="WH/Stock", usage="internal", is_active=True)
    db.add_all([variant, location])
    db.flush()
    db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=5))
    sale = models.SaleOrder(
        reference=reference,
        client_name="Client atelier",
        status=sale_status,
        workflow_type=workflow_type,
        tax_rate=20,
    )
    db.add(sale)
    db.flush()
    visual_config = None
    if with_visual_config:
        visual_config = json.dumps(
            {
                "type": "Fenêtre",
                "width": 1200,
                "height": 1400,
                "material": "ALU",
            }
        )
    db.add(
        models.SaleOrderLine(
            order_id=sale.id,
            description="Menuiserie ALU Sepalumic",
            quantity=1,
            unit_price=1000,
            visual_config=visual_config,
        )
    )
    db.commit()
    return sale.id


def _make_test_client():
    engine = create_engine(
        "sqlite:///:memory:",
        connect_args={"check_same_thread": False},
        poolclass=StaticPool,
    )
    TestingSessionLocal = sessionmaker(autocommit=False, autoflush=False, bind=engine)
    models.Base.metadata.create_all(bind=engine)

    def override_get_db():
        db = TestingSessionLocal()
        try:
            yield db
        finally:
            db.close()

    app.dependency_overrides[database.get_db] = override_get_db
    return engine, TestingSessionLocal, TestClient(app)


def _cleanup_test_client(engine):
    app.dependency_overrides.pop(database.get_db, None)
    models.Base.metadata.drop_all(bind=engine)


SEPALUMIC_CONTENT = b"SEPALUMIC GAMME BASE\r\nVER TEST\r\nRAL;7007;BAVETTE DE FAITAGE;3;barre  6,50\r\n"


def _attach_validated_technical_dossier(
    db,
    sale_id,
    *,
    stock_status="VALIDATED",
    launch_status="TO_REVIEW",
):
    sale = db.query(models.SaleOrder).filter(models.SaleOrder.id == sale_id).one()
    sale.signed_at = datetime(2026, 7, 1, 10, 0)
    client = models.Client(name=f"Client technique {sale.reference}")
    db.add(client)
    db.flush()
    mission = models.MeasureMission(
        reference=f"MET-{sale.reference}",
        client_id=client.id,
        sale_order_id=sale.id,
        status="QUOTED",
        created_by="test",
    )
    dossier = models.TechnicalDossier(
        reference=f"DT-{sale.reference}",
        mission=mission,
        production_status="VALIDATED",
        stock_status=stock_status,
        launch_status=launch_status,
        external_source_system="PROGES",
        external_project_reference="VER TEST",
        created_by="test",
    )
    db.add(dossier)
    db.flush()
    fabrication = models.TechnicalDossierVersion(
        dossier=dossier,
        version_number=1,
        document_type="FABRICATION",
        source_system="PROGES",
        source_reference="VER TEST",
        original_filename="fabrication.pdf",
        stored_filename="fabrication.pdf",
        file_path="uploads/test/fabrication.pdf",
        checksum_sha256=f"fab-{sale.id}",
        opening_ids=[],
        analysis_status="PARSED",
        parsed_summary={},
        parsed_records=[],
        parsed_issues=[],
        created_by="test",
    )
    cutting = models.TechnicalDossierVersion(
        dossier=dossier,
        version_number=2,
        document_type="CUTTING",
        source_system="PROGES",
        source_reference="VER TEST",
        original_filename="SEPVER.TXT",
        stored_filename="SEPVER.TXT",
        file_path="uploads/test/SEPVER.TXT",
        checksum_sha256=f"cut-{sale.id}",
        opening_ids=[],
        analysis_status="PARSED",
        detected_document_type="CUTTING",
        detected_source_system="PROGES",
        detected_project_reference="VER TEST",
        parsed_summary={"record_count": 1},
        parsed_records=[
            {
                "source": "SEPVER.TXT",
                "supplier": "SEPALUMIC",
                "reference": "7007",
                "designation": "Bavette de faitage",
                "quantity": 3,
                "unit": "barre",
                "project_reference": "VER TEST",
            }
        ],
        parsed_issues=[],
        stock_data_approved_at=datetime(2026, 7, 2, 10, 0),
        stock_data_approved_by="chef-stock",
        created_by="test",
    )
    db.add_all([fabrication, cutting])
    db.commit()
    return mission.id, dossier.id, cutting.id


def test_workshop_reservation_moves_sale_to_ready_for_prod():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(db, "VALIDATED", "DEV-READY-AUTO")

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", SEPALUMIC_CONTENT, "text/plain"))],
        )
        assert response.status_code == 200, response.text
        assert response.json()["status"] == "reserved"

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
        assert sale_db.status == "READY_FOR_PROD"
    finally:
        _cleanup_test_client(engine)


def test_workshop_reservation_never_downgrades_advanced_sale_status():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(db, "IN_PRODUCTION", "DEV-DEJA-AVANCE")

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", SEPALUMIC_CONTENT, "text/plain"))],
        )
        assert response.status_code == 200, response.text

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
        assert sale_db.status == "IN_PRODUCTION"
    finally:
        _cleanup_test_client(engine)


def test_workshop_reservation_without_sale_context_does_not_touch_sales():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            product = models.Product(
                reference_base="SEPALUMIC:7007",
                name="Bavette de faitage",
                material_type="ALU",
                unit="barre",
                supplier="SEPALUMIC",
                product_type="stockable",
            )
            db.add(product)
            db.flush()
            variant = models.ProductVariant(
                product_id=product.id,
                reference="SEPALUMIC:7007",
                supplier_reference="7007",
                quantity_in_stock=5,
                min_threshold=0,
            )
            location = models.StockLocation(name="WH/Stock", usage="internal", is_active=True)
            db.add_all([variant, location])
            db.flush()
            db.add(models.StockQuant(variant_id=variant.id, location_id=location.id, quantity=5))
            order = models.Order(
                reference="CMD-DEBIT-LIBRE",
                width=1200,
                height=900,
                material=models.MaterialType.ALU,
                client_name="Client atelier",
                quantity=1,
            )
            db.add(order)
            db.commit()
            order_id = order.id

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"production_order_id": str(order_id)},
            files=[("files", ("SEPVER.TXT", SEPALUMIC_CONTENT, "text/plain"))],
        )
        assert response.status_code == 200, response.text
        reservation = response.json()
        assert reservation["status"] == "reserved"
        assert reservation["sale_order_id"] is None
        assert reservation["production_order_id"] == order_id

        with TestingSessionLocal() as db:
            assert db.query(models.SaleOrder).count() == 0
    finally:
        _cleanup_test_client(engine)


def test_full_flow_reservation_then_launch_production():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-FLOW-COMPLET",
                workflow_type="FABRICATION_FROM_MEASURE",
                with_visual_config=True,
            )

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        reserve_response = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", SEPALUMIC_CONTENT, "text/plain"))],
        )
        assert reserve_response.status_code == 200, reserve_response.text

        with TestingSessionLocal() as db:
            assert db.query(models.SaleOrder).one().status == "READY_FOR_PROD"

        launch_response = client.post(f"/v2/sales/{sale_id}/launch-production", headers=headers)
        assert launch_response.status_code == 200, launch_response.text
        assert launch_response.json()["created_orders"] == 1
        assert launch_response.json()["linked_reservations"] == 1

        with TestingSessionLocal() as db:
            sale_db = db.query(models.SaleOrder).one()
            order = db.query(models.Order).one()
            reservation_db = db.query(models.StockReservation).one()

        assert sale_db.status == "IN_PRODUCTION"
        assert order.sale_order_id == sale_id
        assert reservation_db.production_order_id == order.id
    finally:
        _cleanup_test_client(engine)


def test_technical_cutting_creates_one_traced_reservation_without_reupload():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-CUTTING-AUTO",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, _dossier_id, cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        generic = client.post(
            "/v2/stock/workshop-debits/reservations",
            headers=headers,
            data={"sale_order_id": str(sale_id)},
            files=[("files", ("SEPVER.TXT", SEPALUMIC_CONTENT, "text/plain"))],
        )
        assert generic.status_code == 400
        assert "dossier technique" in generic.text

        legacy_bom = client.post(
            f"/v2/stock/import-bom/{sale_id}",
            headers=headers,
            files={"file": ("bom.csv", b"reference,quantity\n7007,3", "text/csv")},
        )
        assert legacy_bom.status_code == 409
        assert "débit validé" in legacy_bom.text

        first = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert first.status_code == 200, first.text
        reservation = first.json()["execution"]["reservation"]
        assert reservation["status"] == "reserved"
        assert reservation["cutting_version_id"] == cutting_id

        second = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert second.status_code == 200, second.text
        assert second.json()["execution"]["reservation"]["id"] == reservation["id"]

        with TestingSessionLocal() as db:
            reservations = db.query(models.StockReservation).all()
            sale = db.query(models.SaleOrder).filter(models.SaleOrder.id == sale_id).one()
        assert len(reservations) == 1
        assert reservations[0].technical_dossier_version_id == cutting_id
        assert f"id={cutting_id}" in reservations[0].notes
        assert "sha256=cut-" in reservations[0].notes
        assert sale.status == "READY_FOR_PROD"
    finally:
        _cleanup_test_client(engine)


def test_cancelled_technical_reservation_can_be_reactivated_without_duplicate():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-CUTTING-REACTIVATE",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, _dossier_id, cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        created = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert created.status_code == 200, created.text
        reservation = created.json()["execution"]["reservation"]

        cancelled = client.post(
            f"/v2/stock/workshop-debits/reservations/{reservation['id']}/cancel",
            headers=headers,
        )
        assert cancelled.status_code == 200, cancelled.text
        assert cancelled.json()["released_quantity"] > 0

        reactivated = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert reactivated.status_code == 200, reactivated.text
        reactivated_reservation = reactivated.json()["execution"]["reservation"]
        assert reactivated_reservation["id"] == reservation["id"]
        assert reactivated_reservation["status"] == "reserved"
        assert reactivated_reservation["cutting_version_id"] == cutting_id

        with TestingSessionLocal() as db:
            reservations = db.query(models.StockReservation).all()
            lines = db.query(models.StockReservationLine).all()
        assert len(reservations) == 1
        assert all(line.status == "reserved" for line in lines)
        assert all(line.reserved_quantity == line.requested_quantity for line in lines)
    finally:
        _cleanup_test_client(engine)


def test_technical_physical_moves_wait_for_current_launch_authorization():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-GATES-PHYSIQUES",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, dossier_id, _cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        governance = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        ).json()
        reservation = governance["execution"]["reservation"]
        preparation_response = client.post(
            "/v2/stock/workshop-preparations",
            headers=headers,
            json={"reservation_id": reservation["id"]},
        )
        assert preparation_response.status_code == 200, preparation_response.text
        preparation = preparation_response.json()
        for line in preparation["lines"]:
            response = client.patch(
                f"/v2/stock/workshop-preparations/{preparation['id']}/lines/{line['id']}",
                headers=headers,
                json={"prepared_quantity": line["planned_quantity"]},
            )
            assert response.status_code == 200, response.text

        blocked = client.post(
            f"/v2/stock/workshop-preparations/{preparation['id']}/handover",
            headers=headers,
        )
        assert blocked.status_code == 400
        assert "autoriser le lancement" in blocked.text

        with TestingSessionLocal() as db:
            dossier = (
                db.query(models.TechnicalDossier)
                .filter(models.TechnicalDossier.id == dossier_id)
                .one()
            )
            dossier.launched_at = datetime(2026, 7, 3, 10, 0)
            dossier.launch_status = "CORRECTION_REQUIRED"
            db.commit()

        still_blocked = client.post(
            f"/v2/stock/workshop-preparations/{preparation['id']}/handover",
            headers=headers,
        )
        assert still_blocked.status_code == 400
        assert "autoriser le lancement" in still_blocked.text

        authorized = client.patch(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/gate-review",
            headers=headers,
            json={"gate": "LAUNCH", "action": "VALIDATE"},
        )
        assert authorized.status_code == 200, authorized.text

        handed_over = client.post(
            f"/v2/stock/workshop-preparations/{preparation['id']}/handover",
            headers=headers,
        )
        assert handed_over.status_code == 200, handed_over.text
        consumed = client.post(
            f"/v2/stock/workshop-debits/reservations/{reservation['id']}/consume",
            headers=headers,
        )
        assert consumed.status_code == 200, consumed.text
        assert consumed.json()["consumed_lines"] == 1

        duplicate = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert duplicate.status_code == 409
        assert "déjà une réservation consumed" in duplicate.text

        with TestingSessionLocal() as db:
            dossier = (
                db.query(models.TechnicalDossier)
                .filter(models.TechnicalDossier.id == dossier_id)
                .one()
            )
            previous_cutting = (
                db.query(models.TechnicalDossierVersion)
                .filter(models.TechnicalDossierVersion.document_type == "CUTTING")
                .one()
            )
            db.add(
                models.TechnicalDossierVersion(
                    dossier=dossier,
                    version_number=3,
                    document_type="CUTTING",
                    source_system="PROGES",
                    source_reference="VER TEST",
                    original_filename="SEPVER-V2.TXT",
                    stored_filename="SEPVER-V2.TXT",
                    file_path="uploads/test/SEPVER-V2.TXT",
                    checksum_sha256="cut-after-consumption",
                    opening_ids=[],
                    analysis_status="PARSED",
                    parsed_summary={"record_count": 1},
                    parsed_records=previous_cutting.parsed_records,
                    parsed_issues=[],
                    stock_data_approved_at=datetime(2026, 7, 5, 10, 0),
                    stock_data_approved_by="chef-stock",
                )
            )
            db.commit()

        revision = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert revision.status_code == 409
        assert "régularisation de stock" in revision.text
        with TestingSessionLocal() as db:
            assert db.query(models.StockReservation).count() == 1
    finally:
        _cleanup_test_client(engine)


def test_launch_gate_rejects_cancelled_or_previous_cutting_reservation():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-GATE-VERSION",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, dossier_id, cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )

        headers = _auth_headers(TestingSessionLocal, "atelier-manager")
        created = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/reservation",
            headers=headers,
        )
        assert created.status_code == 200, created.text
        reservation_id = created.json()["execution"]["reservation"]["id"]

        with TestingSessionLocal() as db:
            reservation = (
                db.query(models.StockReservation)
                .filter(models.StockReservation.id == reservation_id)
                .one()
            )
            reservation.status = "cancelled"
            cutting = (
                db.query(models.TechnicalDossierVersion)
                .filter(models.TechnicalDossierVersion.id == cutting_id)
                .one()
            )
            cutting_v2 = models.TechnicalDossierVersion(
                dossier_id=dossier_id,
                version_number=3,
                document_type="CUTTING",
                source_system="PROGES",
                source_reference="VER TEST",
                original_filename="SEPVER-V2.TXT",
                stored_filename="SEPVER-V2.TXT",
                file_path="uploads/test/SEPVER-V2.TXT",
                checksum_sha256="cut-v2",
                opening_ids=[],
                analysis_status="PARSED",
                parsed_summary={"record_count": 1},
                parsed_records=cutting.parsed_records,
                parsed_issues=[],
                stock_data_approved_at=datetime(2026, 7, 4, 10, 0),
                stock_data_approved_by="chef-stock",
            )
            db.add(cutting_v2)
            db.commit()

        rejected = client.patch(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/gate-review",
            headers=headers,
            json={"gate": "LAUNCH", "action": "VALIDATE"},
        )
        assert rejected.status_code == 409
        assert "dernière version de débit" in rejected.text
    finally:
        _cleanup_test_client(engine)


def test_operator_cannot_mutate_technical_dossier():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-RBAC-TECHNIQUE",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, _dossier_id, _cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )

        operator = _auth_headers(TestingSessionLocal, "operator-technique", role="OPERATOR")
        create = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier",
            headers=operator,
        )
        assert create.status_code == 403

        upload = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/versions",
            headers=operator,
            data={
                "document_type": "CUTTING",
                "source_system": "INTERNAL",
            },
            files={"file": ("debit.txt", SEPALUMIC_CONTENT, "text/plain")},
        )
        assert upload.status_code == 403

        submit = client.patch(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/submit",
            headers=operator,
            params={"phase": "CUTTING"},
        )
        assert submit.status_code == 403

        review = client.patch(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/review",
            headers=operator,
            json={"phase": "CUTTING", "action": "VALIDATE"},
        )
        assert review.status_code == 403
    finally:
        _cleanup_test_client(engine)


def test_internal_supplier_document_does_not_conflict_with_external_design_source():
    engine, TestingSessionLocal, client = _make_test_client()
    try:
        with TestingSessionLocal() as db:
            sale_id = _seed_stock_and_sale(
                db,
                "VALIDATED",
                "DEV-SOURCE-DERIVEE",
                workflow_type="FABRICATION_FROM_MEASURE",
            )
            mission_id, dossier_id, _cutting_id = _attach_validated_technical_dossier(
                db,
                sale_id,
            )
            dossier = (
                db.query(models.TechnicalDossier)
                .filter(models.TechnicalDossier.id == dossier_id)
                .one()
            )
            dossier.production_status = "CORRECTION_REQUIRED"
            db.commit()

        headers = _auth_headers(TestingSessionLocal, "be-manager")
        upload = client.post(
            f"/v2/mmg/missions/{mission_id}/technical-dossier/versions",
            headers=headers,
            data={
                "document_type": "FABRICATION",
                "source_system": "INTERNAL",
                "source_reference": "VER TEST",
            },
            files={
                "file": (
                    "commande-cortizo.txt",
                    b"COMMANDE FOURNISSEUR CORTIZO\nREFERENCE VER TEST\n",
                    "text/plain",
                )
            },
        )
        assert upload.status_code == 200, upload.text
        payload = upload.json()
        latest = payload["versions"][-1]
        assert latest["source_system"] == "INTERNAL"
        assert not any(
            issue.get("code") == "technical_dossier_source_mismatch"
            for issue in latest["parsed_issues"]
        )
        assert payload["external_source_system"] == "PROGES"
    finally:
        _cleanup_test_client(engine)
